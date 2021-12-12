import logging
import os
import sys
import subprocess
import openpyxl

from PyQt5 import uic
from PyQt5.QtWidgets import QMainWindow, QApplication, QFileDialog, QSpacerItem, QSizePolicy, QMessageBox

from core.Core import DateManager, Customer, Employee, Transaction, Product
from core.Widgets import Terminate_Widget, TaskImportWidget
import core.CSSTemplates

class TaskImporter(QMainWindow, DateManager):
    def __init__(self):
        QMainWindow.__init__(self)
        DateManager.__init__(self)
        self.logger = logging.getLogger(self.__class__.__name__)

        self.me = Employee()
        self.me.set_login(os.getlogin())
        self.importedtasks = []
        self.tiw_current = 0

        try:
            self.ui = uic.loadUi(r'src\ui\task_import_mainwindow.ui', self)
        except Exception:
            self.logger.exception(f'Error while trying to find and load ui file')
            Terminate_Widget(20)

        self._setup_ui()
        self._connect_signals()

        self.manual_path = r'src\man\Task Importer - Manual.pdf'

    def _setup_ui(self):
        self.verticalSpacer = QSpacerItem(20, 40, QSizePolicy.Minimum, QSizePolicy.Expanding)
        self.ui.task_containter_vlayout.addItem(self.verticalSpacer)

    def _connect_signals(self):
        self.ui.save_tasks_button.clicked.connect(self._activated_save_tasks_button)
        self.ui.excel_import_act.triggered.connect(self._activated_excel_import_act)
        self.ui.help_manual_act.triggered.connect(self._activated_help_manual_act)
        self.ui.clear_act.triggered.connect(self._clear_window)
        self.ui.add_1_row_act.triggered.connect(self._activated_add_1_row_act)
        self.ui.add_2_row_act.triggered.connect(self._activated_add_2_row_act)
        self.ui.add_3_row_act.triggered.connect(self._activated_add_3_row_act)
        self.ui.add_4_row_act.triggered.connect(self._activated_add_4_row_act)
        self.ui.add_5_row_act.triggered.connect(self._activated_add_5_row_act)
        self.ui.add_10_row_act.triggered.connect(self._activated_add_10_row_act)
        self.ui.add_20_row_act.triggered.connect(self._activated_add_20_row_act)
        self.ui.add_50_row_act.triggered.connect(self._activated_add_50_row_act)

    def _message_receiver(self, msg):
        self.ui.statusBar().showMessage(str(msg))

    def _clear_window(self):
        for i in range(len(self.importedtasks)-1, -1, -1):
            tiw = self.importedtasks[i]
            self.ui.task_containter_vlayout.removeWidget(tiw)
            self.importedtasks.remove(tiw)
            tiw.deleteLater()
            tiw = None
            self.tiw_current -= 1

    def _add_empty_rows(self, rows_count):
        self.ui.task_containter_vlayout.removeItem(self.verticalSpacer)
        for i in range(0, rows_count):
            tiw = TaskImportWidget(self.tiw_current + 1, self.me)
            self.tiw_current += 1
            self.importedtasks.append(tiw)
            self.ui.task_containter_vlayout.addWidget(tiw)
            tiw.signal_customer_pass.connect(self.auto_fill)
            tiw.signal_message.connect(self._message_receiver)
        self.ui.task_containter_vlayout.addItem(self.verticalSpacer)

    def _activated_excel_import_act(self):
        path = QFileDialog.getOpenFileName(caption="Select Excel file with tasks",
                                           directory=r"C:\\",
                                           filter="Excel(*.xls, *.xlsx)")
        if path[0]:
            tasks = self.import_data_from_excel(path[0])
            if tasks:
                self.ui.task_containter_vlayout.removeItem(self.verticalSpacer)
                for task in tasks:
                    tiw = TaskImportWidget(self.tiw_current + 1, self.me)
                    self.tiw_current += 1
                    tiw.load_parameters(task)
                    self.importedtasks.append(tiw)
                    self.ui.task_containter_vlayout.addWidget(tiw)
                    tiw.signal_customer_pass.connect(self.auto_fill)
                    tiw.signal_message.connect(self._message_receiver)
                self.ui.task_containter_vlayout.addItem(self.verticalSpacer)
            else:
                info = QMessageBox.information(self,
                                               'Information',
                                               'There is no data to be imported')

    def _activated_help_manual_act(self):
        try:
            subprocess.Popen([self.manual_path], shell=True)
        except Exception:
            self.logger.exception("Incorrect path for Manual document")

    def _activated_add_1_row_act(self):
        self._add_empty_rows(1)

    def _activated_add_2_row_act(self):
        self._add_empty_rows(2)

    def _activated_add_3_row_act(self):
        self._add_empty_rows(3)

    def _activated_add_4_row_act(self):
        self._add_empty_rows(4)

    def _activated_add_5_row_act(self):
        self._add_empty_rows(5)

    def _activated_add_10_row_act(self):
        self._add_empty_rows(10)

    def _activated_add_20_row_act(self):
        self._add_empty_rows(20)

    def _activated_add_50_row_act(self):
        self._add_empty_rows(50)

    def _activated_save_tasks_button(self):
        for tiw in self.importedtasks:
            if tiw.widget_id.isChecked() and tiw.isEnabled():
                tiw._register_task()
                tiw.setEnabled(False)

    def auto_fill(self, parameters):
        if self.ui.downstream_auto_fill_act.isChecked():
            need_action = self.importedtasks[parameters["widget_id"]:]
            for tiw in need_action:
                tiw.load_parameters(parameters)

    def import_data_from_excel(self, filename):
        filename = r'{}'.format(filename)
        if os.path.isfile(filename):
            book = openpyxl.load_workbook(filename)
            sheets = tuple(book.get_sheet_names())
            to_import_list = []

            if 'TaTO_IMPORT' in sheets:
                to_import_sh = book.get_sheet_by_name('TaTO_IMPORT')
                for row in to_import_sh[f'A3:L{to_import_sh.max_row}']:
                    try:
                        str_customer = str(row[0].value)
                        str_transaction = str(row[1].value)
                        str_discipline = str(row[2].value)
                        str_activity_type = str(row[3].value)
                        str_task_name = str(row[4].value)
                        str_responsible_name = str(row[5].value)
                        str_planned_start = str(row[6].value)
                        str_planned_end = str(row[7].value)
                        str_planned_delivery = str(row[9].value)

                        task = {"customer": str_customer,
                                "transaction": str_transaction,
                                "discipline": str_discipline,
                                "activity_type": str_activity_type,
                                "task_name": str_task_name,
                                "responsible_name": str_responsible_name,
                                "planned_start": str_planned_start,
                                "planned_end": str_planned_end,
                                "planned_delivery": str_planned_delivery}


                        if str(row[8].value).upper() == "YES" or str(row[11].value) == "1" or str(row[11].value).upper() == "TRUE":
                            task["deliverable"] = True
                        if str(row[8].value).upper() == "NO" or str(row[11].value) == "0" or str(row[11].value).upper() == "FALSE":
                            task["deliverable"] = False

                        try:
                            float_planned_hours = float(row[10].value)
                        except Exception:
                            pass
                        else:
                            task["planned_hours"] = float_planned_hours

                        if str(row[11].value).upper() == "YES" or str(row[11].value) == "1" or str(row[11].value).upper() == "TRUE":
                            task["checking_required"] = True

                        if str(row[11].value).upper() == "NO" or str(row[11].value) == "0" or str(row[11].value).upper() == "FALSE":
                            task["checking_required"] = False

                        to_import_list.append(task)
                    except Exception as err:
                        print(err)
                return to_import_list
            else:
                self._message_receiver("Can't find sheet \"TaTO_IMPORT\" in selected file")
                buttonReply = QMessageBox.question(self, 'Improper structure', "Can't find sheet \"TaTO_IMPORT\" in selected file.\nWould you like add this sheet?",
                                                   QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                if buttonReply == QMessageBox.Yes:
                    if self.prepare_excel(filename):
                        try:
                            info = QMessageBox.information(self,
                                                           'Information',
                                                           'Sheet "TaTO_IMPORT" has been added.\nPlease prepare data for import')
                        except Exception as err:
                            warn = QMessageBox.warning(self,
                                                       'Warning',
                                                       "Operation not possible")
                    else:
                        pass
                else:
                    pass

    def prepare_excel(self, filename):
        try:
            filename = r'{}'.format(filename)
            if os.path.isfile(filename):
                book = openpyxl.load_workbook(filename)
                sheets = tuple(book.get_sheet_names())

                if not 'TaTO_IMPORT' in sheets:
                    book.create_sheet("TaTO_IMPORT")
                    tato_sheet = book.get_sheet_by_name('TaTO_IMPORT')
                    tato_sheet.merge_cells('A1:L1')
                    tato_sheet["A1"] = """ARKUSZ TEN ( I JEGO STRUKTURA) ZOSAŁY ZAPROJEKTOWANE TAK ABY DANE W NIM ZAWARTE MOŻNA BYŁO IMPORTOWAĆ DO APLIKACJI TaTO. 
W CELU IMPORTU NALEŻY WYPELNIĆ ARKUSZ ODPOWIEDNIMI DANYMI I POSTĘPOWAC ZGODNIE Z INTRUKCJĄ DOSTĘPNĄ W TaTO\n
UWAGA! NIE NALEŻY ZMIENIAĆ NAZWY TEGO ARKUSZA ANI DOKONYWAĆ ŻADNYCH MODYFIKACJI BIERZĄCEJ  STRUKTURY, W PRZECIWNYM RAZIE DANE MOGĄ IMPORTOWAĆ SIĘ NIEPRAWODŁOWO"""
                    tato_sheet["A2"] = "customer"
                    tato_sheet["B2"] = "transaction"
                    tato_sheet["C2"] = "discipline"
                    tato_sheet["D2"] = "activity"
                    tato_sheet["E2"] = "task name"
                    tato_sheet["F2"] = "responsible"
                    tato_sheet["G2"] = "planned start"
                    tato_sheet["H2"] = "planned end"
                    tato_sheet["I2"] = "deliverable"
                    tato_sheet["J2"] = "planned delivery"
                    tato_sheet["K2"] = "planned hours"
                    tato_sheet["L2"] = "checker required"

                    tato_sheet.column_dimensions['A'].width = 20
                    tato_sheet.column_dimensions['B'].width = 20
                    tato_sheet.column_dimensions['C'].width = 20
                    tato_sheet.column_dimensions['D'].width = 20
                    tato_sheet.column_dimensions['E'].width = 50
                    tato_sheet.column_dimensions['F'].width = 20
                    tato_sheet.column_dimensions['G'].width = 15
                    tato_sheet.column_dimensions['H'].width = 15
                    tato_sheet.column_dimensions['I'].width = 15
                    tato_sheet.column_dimensions['J'].width = 15
                    tato_sheet.column_dimensions['K'].width = 15
                    tato_sheet.column_dimensions['L'].width = 15

                    tato_sheet.row_dimensions[1].height = 50
                    tato_sheet['A1'].alignment = openpyxl.styles.Alignment(wrap_text=True)

                    book.save(filename)
        except Exception:
            return False
        else:
            return True


if __name__ == "__main__":
    os.chdir(os.getcwd() + "\..")
    app = QApplication(sys.argv)
    # app.setStyleSheet(core.CSSTemplates.my_style)
    ti = TaskImporter()
    ti.ui.show()
    sys.exit(app.exec_())
