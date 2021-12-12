#THIS DOCUMENT DOES NOT CONTAIN ANY TECHNICAL DATA

import sqlite3
import logging
import configparser
import os
import sys
import openpyxl
import copy
import numpy as np
import win32com.client as win32
import time
from PyQt5.QtWidgets import QWidget, QDateEdit, QVBoxLayout, QMessageBox, QFileDialog
from PyQt5.QtCore import pyqtSignal, QObject, QDate, QDateTime
from datetime import datetime, timedelta, date

# information
"""
 /$$$$$$$$   /$$$$$$   /$$$$$$$$   /$$$$$$ 
|__  $$__/  /$$__  $$ |__  $$__/  /$$__  $$
   | $$    | $$  \ $$    | $$    | $$  \ $$
   | $$    | $$$$$$$$    | $$    | $$  | $$
   | $$    | $$__  $$    | $$    | $$  | $$
   | $$    | $$  | $$    | $$    | $$  | $$
   | $$ /$$| $$  | $$ /$$| $$ /$$|  $$$$$$/
   |__/|__/|__/  |__/|__/|__/|__/ \______/ 

Module:         Core
Version         1.0 alpha 1
Release date:   18.01.2019

This module is integral part of application T.A.T.O 
It CAN'T be used whitout colaboration with other T.A.T.O core.
"""

import pyautogui as py


class Status:
    def __init__(self):
        self.code = {
                "READY TO POP": 10,
                "ASSIGNED": 15,
                "SCHEDULED": 20,
                "ONGOING": 40,
                "LATE": 45,
                "READY TO CHECK": 50,
                "IN CHECK": 60,
                "REWORK NEEDED": 90,
                "IN REWORK": 100,
                "FINISHED": 110,
                "RETURNED": 112,
                "RELEASED": 115,
                "NOT READY TO DELIVERY": 117,
                "READY TO DELIVERY": 120,
                "DELIVERED": 130,
                "ERROR": 666,
                }

    def name(self, code_value):
            for key, value in self.code.items():
                    if value == code_value:
                            return key
            return "CODE ERROR"

status = Status()

class Config:
    def __init__(self, config_directory):
        logger = logging.getLogger(__name__ + ".Config")
        logger.info(f'create configuration object')

        self.main_config_path = config_directory + 'TATO.config'
        logger.debug(f'path for config file: {self.main_config_path}')
        self.configParser = configparser.ConfigParser()

        if os.path.isfile(self.main_config_path):
            self.configParser.read(config_directory + 'PyWRM.config')
            logger.info(f'reading of the configuration file -> started')
            try:
                self.db_path = self.configParser.get('path_config', 'db_path')
            except Exception as err:
                logger.error("Exception occurred")
                sys.exit(1)
            logger.info(f'reading of the configuration file -> finished')
        else:
            logger.error(f"can't find configuration file at {self.main_config_path}")


class PersonalConfig:
    def __init__(self):
        self.logger = logging.getLogger(self.__class__.__name__)
        self.logger.debug(f'crating object')
        self.config_dir_path = os.getcwd() + '\\config\\'
        self.logger.debug(f'config directory path: {self.config_dir_path}')
        self.my_login = os.getlogin()
        self.logger.debug(f'login: {self.my_login}')

        self.changes = False

        if os.path.isfile(self.config_dir_path + self.my_login + '.config'):
            self.my_config = self.config_dir_path + self.my_login + '.config'
        else:
            self.my_config = self.config_dir_path + 'default.config'
        self.logger.info(f'personal settings taken from: {self.my_config}')

        self.config = configparser.ConfigParser()
        self.config.read(self.my_config)
        self.logger.info(f'reading configuration settings')
        try:
            self.task_manager_show_task_info = bool(int(self.config.get('task_manager', 'show_task_info')))
        except Exception:
            self.logger.exception(f'Missing setting in config file.\nError occured')
            sys.exit(5)

        try:
            self.task_manager_show_transaction_info = bool(int(self.config.get('task_manager', 'show_transaction_info')))
        except Exception:
            self.logger.exception(f'Missing setting in config file.\nError occured')
            sys.exit(5)

        try:
            self.task_manager_show_workload_preview = bool(int(self.config.get('task_manager', 'show_workload_preview')))
        except Exception:
            self.logger.exception(f'Missing setting in config file: show_workload_preview')
            sys.exit(5)

        try:
            self.task_manager_workload_preview_days = int(self.config.get('task_manager', 'workload_preview_days'))
        except Exception:
            self.logger.exception(f'Missing setting in config file: workload_preview_days')

    def set_task_manager_show_task_info(self, val):
        if val != self.task_manager_show_task_info:
            self.task_manager_show_task_info = val
            self.changes = True

    def set_task_manager_show_transaction_info(self, val):
        if val != self.task_manager_show_transaction_info:
            self.task_manager_show_transaction_info = val
            self.changes = True

    def set_task_manager_show_workload_preview(self, val):
        if val != self.task_manager_show_workload_preview:
            self.task_manager_show_workload_preview = val
            self.changes = True

    def save(self):
        if self.changes:
            self.config.set('task_manager', 'show_task_info', str(int(self.task_manager_show_task_info)))
            self.config.set('task_manager', 'show_transaction_info', str(int(self.task_manager_show_transaction_info)))
            self.config.set('task_manager', 'show_workload_preview', str(int(self.task_manager_show_workload_preview)))

            with open(self.config_dir_path + self.my_login + '.config', 'w') as file:
                self.config.write(file)


class ConnectionDB(QObject):
    signal_exporting_status = pyqtSignal(str)
    signal_operations = pyqtSignal(str)

    def __init__(self, db_path=None, memory_dump=False):
        QObject.__init__(self)
        self.logger = logging.getLogger(self.__class__.__name__)
        if db_path:
            self.db_path = db_path
        else:
            # self.db_path = os.getcwd() + '\\db\\database.db'
            self.db_path = r'G:\DR\Power and Controls\04. Administration\06. TATO\db\database.db'

        if self.db_avaliable():
            self.connection = sqlite3.connect(self.db_path)
            self.cursor = self.connection.cursor()
            if memory_dump:
                self.memory_dump()
            self.logger.info(f'connection successfull')
            print(f'Connected to {self.db_path}')
        else:
            self.logger.error(f'CONNECTION FAILED')
            sys.exit(4)



    def memory_dump(self):
        print("dumping")
        disk_connection = self.connection
        self.connection = sqlite3.connect(':memory:')
        self.cursor = self.connection.cursor()
        query = "".join(line for line in disk_connection.iterdump())
        self.connection.executescript(query)


    def execute(self, cmd, param=None):
        try:
            if param:
                self.cursor.execute(cmd, param)
            else:
                self.cursor.execute(cmd)
        except Exception as err:
            print(err)
            self.signal_operations.emit("Unable to execute SQL Command. Check connections")


    def commit(self):
        try:
            self.connection.commit()
        except Exception as err:
            print(err)
            self.signal_operations.emit("Unable to save. Check connection")


    def fetchall(self):
        try:
            return self.cursor.fetchall()
        except Exception as err:
            print(err)


    def fetchone(self):
        try:
            return self.cursor.fetchone()
        except Exception as err:
            print(err)


    def db_avaliable(self):
        if os.path.isfile(self.db_path):
            return True
        else:
            return False


    def export_data_to_excel(self, filename):
        filename = r'{}'.format(filename)
        if os.path.isfile(filename):
            book = openpyxl.load_workbook(filename)
            sheets = tuple(book.get_sheet_names())

            if 'task_data' and 'transaction_data' and 'employee_data' and 'TASKS TO ANALYZE' in sheets:
                # SAVING TASKS
                self.signal_exporting_status.emit("Downloading tasks data")
                sh_tasks = book.get_sheet_by_name('task_data')
                for row in sh_tasks['A2:T10000']:
                    for cell in row:
                        cell.value = None

                self.cursor.execute("SELECT * FROM tasks")
                all_tasks = self.fetchall()

                for row in range(0, len(all_tasks)):
                    for col in range(0, len(all_tasks[row])):
                        sh_tasks.cell(row=row + 2, column=col + 1).value = all_tasks[row][col]

                # SAVING TRANSACTIONS
                self.signal_exporting_status.emit("Downloading transactions data")
                sh_transactions = book.get_sheet_by_name('transaction_data')
                for row in sh_transactions['A2:Q10000']:
                    for cell in row:
                        cell.value = None

                self.cursor.execute("SELECT * FROM transactions")
                all_transactions = self.fetchall()

                for row in range(0, len(all_transactions)):
                    for col in range(0, len(all_transactions[row])):
                        sh_transactions.cell(row=row + 2, column=col + 1).value = all_transactions[row][col]

                # SAVING EMPLOYEES
                self.signal_exporting_status.emit("Downloading employees data")
                sh_employees = book.get_sheet_by_name('employee_data')
                for row in sh_employees['A2:P1000']:
                    for cell in row:
                        cell.value = None

                self.cursor.execute("SELECT * FROM employees")
                all_employees = self.fetchall()

                for row in range(0, len(all_employees)):
                    for col in range(0, len(all_employees[row])):
                        sh_employees.cell(row=row + 2, column=col + 1).value = all_employees[row][col]

                # UPDATING DOWNLOAD TIME
                sh_task_analyze = book.get_sheet_by_name('TASKS TO ANALYZE')
                sh_task_analyze['A1'] = 'downloaded: {}'.format(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

            self.signal_exporting_status.emit("Saving data to Excel file")
            try:
                book.save(filename)
            except PermissionError:
                self.signal_exporting_status.emit("Export FAILED. Can't save file. Excel file must be closed while exporting")
            else:
                self.signal_exporting_status.emit("Export SUCCESSFULL")


    def __del__(self):
        try:
            self.connection.close()
        except AttributeError:
            pass


class DateManager:

    def __init__(self):
        self.format_for_date = '%Y-%m-%d'
        self.format_for_widget = 'yyyy-MM-dd'

    def convert_to_date_class(self, date_str, trimmed=True):
        if isinstance(date_str, str):
            if trimmed:
                if len(date_str) > 9:
                    date_only = date_str.split(' ')[0]
                    return datetime.strptime(date_only, self.format_for_date).date()
                else:
                    return None
            else:
                return datetime.strptime(date_str, self.format_for_date).date()
        elif isinstance(date_str, date):
            return date_str
        else:
            return None

    def convert_to_string_class(self, date_class):
        if date_class:
            return date_class.strftime(self.format_for_date)
        else:
            return None

    def daterange(self, start_date, end_date):
        for n in range(int((end_date - start_date).days + 1)):
            yield start_date + timedelta(n)

    def get_first_day(self, dt, d_years=0, d_months=0):
        # d_years, d_months are "deltas" to apply to dt
        y, m = dt.year + d_years, dt.month + d_months
        a, m = divmod(m - 1, 12)
        return date(y + a, m + 1, 1)

    def get_last_day(self, dt):
        return self.get_first_day(dt, 0, 1) + timedelta(-1)

    def convert_pyDateTime_to_QDate(self, pyDateTime):
        str_date = self.convert_to_string_class(pyDateTime)
        qDate = QDate.fromString(str_date,self.format_for_widget)
        return qDate

    def convert_QDate_to_pyDate(self, QDate):
        pyDate = QDate.toPyDate()
        return pyDate


class EmailSender(QObject, DateManager):
    def __init__(self):
        self.logger = logging.getLogger(self.__class__.__name__)
        QObject.__init__(self)
        DateManager.__init__(self)
        outlook = win32.Dispatch('outlook.application')
        self.mail = outlook.CreateItem(0)

    def new_task_notification(self, task):
        try:
            link_to_tato_bat = 'G:/DR/Power and Controls/04. Administration/06. TATO/START TATO.bat'
            if task.deliverable == 1:
                delivery_date = task.planned_delivery
            else:
                delivery_date = "non deliverable to customer"
            self.mail.To = task.responsible.e_mail
            self.mail.Subject = 'T.A.T.O. - New Task Notification'
            self.mail.HTMLBody = f'''<p>Dear {task.responsible.full_name},</p>
                                    <p>you have been notified because <strong>new task</strong> was assigned to you in T.A.T.O. software.</p>
                                    <p>DETAILS:</p>
                                    <table border-color: black; float: left;" border="1">
                                    <tbody>
                                    <tr>
                                    <td style="width: 128px;">Task</td>
                                    <td style="width: 420px;">{task.name}</td>
                                    </tr>
                                    <tr>
                                    <td style="width: 128px;">Creator</td>
                                    <td style="width: 180px;">{task.creator.full_name}</td>
                                    </tr>
                                    <tr>
                                    <td style="width: 128px;">Responsible</td>
                                    <td style="width: 180px;">{task.responsible.full_name}</td>
                                    </tr>
                                    <tr>
                                    <td style="width: 128px;">Planned start</td>
                                    <td style="width: 180px;">{task.planned_start}</td>
                                    </tr>
                                    <tr>
                                    <td style="width: 128px;">Planned end</td>
                                    <td style="width: 180px;">{task.planned_end}</td>
                                    </tr>
                                    <tr>
                                    <td style="width: 128px;">Planned delivery</td>
                                    <td style="width: 180px;">{task.planned_end}</td>
                                    </tr>
                                    <tr>
                                    <td style="width: 128px;">Planned hours&nbsp;</td>
                                    <td style="width: 180px;">{task.hours_planned}</td>
                                    </tr>
                                    </tbody>
                                    </table>
                                    <p>To manage task activities please run <a href="{link_to_tato_bat}">T.A.T.O.</a> software</p>'''
        except Exception:
            self.logger.exception(F'Error while trying to prepare message for task')

    def show_mail(self):
        self.mail.Display()

    def send_mail(self):
        try:
            self.mail.Send()
        except Exception:
            self.logger.exception(f'Error while trying to send e-mail notification')


class Customer(QObject, DateManager):
    changed_id = pyqtSignal(int, name='customer_id_changed')
    ram_customers = dict()

    def __init__(self, db=None):
        self.logger = logging.getLogger(self.__class__.__name__)
        self.logger.debug(f'crating object')
        QObject.__init__(self)
        DateManager.__init__(self)
        if db:
            self.db = db
        else:
            self.logger.warning(f'connecting to the database using NEW HANDLER')
            self.db = ConnectionDB()

        #SQL TABLE
        self.id = None
        self.name = None
        #LISTS
        self.names = []
        self.ids = []

        self.transactions = []
        self.products = []

        self.statistics = {'tasks_total_d': [],
                           'tasks_on_time_d': [],
                           'tasks_total_nd': [],
                           'tasks_on_time_nd': [],
                           'transactions_total_d': [],
                           'transactions_on_time_d': [],
                           'transactions_total_nd': [],
                           'transactions_on_time_nd': []}

    def calculate_statistics(self):
        self._calculate_statistics_task_deliverable()
        self._calculate_statistics_task_nondeliverable()
        self._calculate_statistics_transactions_deliverable()
        self._calculate_statistics_transactions_nondeliverable()

    def _calculate_statistics_task_deliverable(self):
        cmd = '''SELECT
                    id,
                    planned_delivery,
                    true_delivery
                FROM
                    tasks
                WHERE
                    transaction_id
                IN(
                    SELECT
                        id
                    FROM
                        transactions
                    WHERE
                        customer_id = ?)
                AND
                    deliverable = 1'''
        param = [self.id, ]
        try:
            self.db.execute(cmd, param)
        except Exception:
            self.logger.exception(f'Error while trying execute command on database')
        else:
            try:
                sql_res = self.db.fetchall()
            except Exception:
                self.logger.exception(f'Error while trying fetching data from database')
            else:
                self.statistics['tasks_on_time_d'], self.statistics['tasks_total_d'] = self._compute_dates(sql_res)

    def _calculate_statistics_task_nondeliverable(self):
        cmd = '''SELECT
                    id,
                    planned_end,
                    true_end
                FROM
                    tasks
                WHERE
                    transaction_id
                IN(
                    SELECT
                        id
                    FROM
                        transactions
                    WHERE
                        customer_id = ?)
                AND
                    deliverable = 0'''
        param = [self.id, ]
        try:
            self.db.execute(cmd, param)
        except Exception:
            self.logger.exception(f'Error while trying execute command on database')
        else:
            try:
                sql_res = self.db.fetchall()
            except Exception:
                self.logger.exception(f'Error while trying fetching data from database')
            else:
                self.statistics['tasks_on_time_nd'], self.statistics['tasks_total_nd'] = self._compute_dates(sql_res)

    def _calculate_statistics_transactions_deliverable(self):
        cmd = '''SELECT
                    id,
                    planned_delivery,
                    true_delivery
                FROM
                    transactions
                WHERE
                    customer_id = ?
                AND
                    deliverable = 1'''

        param = [self.id, ]
        try:
            self.db.execute(cmd, param)
        except Exception:
            self.logger.exception(f'Error while trying execute command on database')
        else:
            try:
                sql_res = self.db.fetchall()
            except Exception:
                self.logger.exception(f'Error while trying fetching data from database')
            else:
                self.statistics['transactions_on_time_d'], self.statistics['transactions_total_d'] = self._compute_dates(sql_res)

    def _calculate_statistics_transactions_nondeliverable(self):
        cmd = '''SELECT
                    id,
                    planned_end,
                    true_end
                FROM
                    transactions
                WHERE
                    customer_id = ?
                AND
                    deliverable = 0'''

        param = [self.id, ]
        try:
            self.db.execute(cmd, param)
        except Exception:
            self.logger.exception(f'Error while trying execute command on database')
        else:
            try:
                sql_res = self.db.fetchall()
            except Exception:
                self.logger.exception(f'Error while trying fetching data from database')
            else:
                self.statistics['transactions_on_time_nd'], self.statistics['transactions_total_nd'] = self._compute_dates(sql_res)

    def _compute_dates(self, sql_list):
        on_time = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
        total = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
        for element in sql_list:
            id = element[0]
            planned_date = self.convert_to_date_class(element[1])
            true_date = self.convert_to_date_class(element[2])
            if planned_date.year == 2019:
                if planned_date.month == 12:
                    oko = 1
            # if planned_date.year == datetime.now().date().year:
                if true_date is not None:
                    if planned_date >= true_date:
                        on_time[planned_date.month - 1] += 1
                    total[planned_date.month - 1] += 1
                else:
                    if planned_date <= datetime.now().date():
                        total[planned_date.month - 1] += 1
        return np.array(on_time), np.array(total)

    def reset(self):
        self.logger.debug(f'reseting atributes')
        self.id = None
        self.name = None
        self.loaded = False
        self.transactions.clear()
        self.products.clear()

    def set_id(self, id):
        if self.id != id:
            self.reset()
            self.id = id
            self.load_data()
            self.changed_id.emit(id)

    def set_id_by_index(self, index):
        self.set_id(self.ids[index])

    def load_names(self):
        self.logger.info(f'loading names')
        self.ids.clear()
        self.names.clear()
        self.ids.append(None)
        self.names.append(" ")
        cmd = 'SELECT id, name FROM customers'
        param = []
        try:
            self.db.execute(cmd, param)
        except Exception as err:
            self.logger.exception(f'FAILED loading names.\nError occured')
        else:
            for element in self.db.fetchall():
                self.ids.append(element[0])
                self.names.append(element[1])

    def load_data(self):
        if self.id in self.ram_customers.keys():
            print('(CUSTOMER) loading data from RAM')
            self.name = self.ram_customers[self.id].name
        else:
            customer_data = Customer(db=self.db)
            if self.id is not None:
                print('(CUSTOMER) loading data from DATABASE')
                cmd = '''SELECT
                            name
                        FROM
                            customers
                        WHERE
                            id = ?'''
                param = [self.id, ]
                try:
                    self.db.execute(cmd, param)
                except Exception:
                    self.logger.exception(f'Error while trying execute command on database')
                else:
                    try:
                        sql_res = self.db.fetchone()
                    except Exception:
                        self.logger.exception(f'Error while trying to fetching data from database')
                    else:
                        self.name = sql_res[0]

                        customer_data.id = self.id
                        customer_data.name = self.name
                        self.ram_customers[self.id] = customer_data

    def load_products(self):
        if self.id is not None:
            cmd = '''SELECT
                        product_id
                    FROM
                        customer_product
                    WHERE
                        customer_id = ?'''
            param = [self.id, ]
            try:
                self.db.execute(cmd, param)
            except Exception:
                self.logger.exception(f'Error whie trying execute command on database')
            else:
                try:
                    sql_res = self.db.fetchall()
                except Exception:
                    self.loger.exception(f'Error while trying to fetching data from database')
                else:
                    self.products.clear()
                    for product_id in sql_res:
                        product = Product(db=self.db)
                        product.set_id(product_id[0])
                        self.products.append(product)

    def load_transactions(self):
        if self.id is not None:
            cmd = '''SELECT
                        id
                    FROM
                        transactions
                    WHERE
                        customer_id = ?
                        '''
            param = [self.id, ]
            try:
                self.db.execute(cmd, param)
            except Exception:
                self.logger.exception(f'Error whie trying execute command on database')
            else:
                try:
                    sql_res = self.db.fetchall()
                except Exception:
                    self.loger.exception(f'Error while trying to fetching data from database')
                else:
                    self.transactions.clear()
                    for transaction_id in sql_res:
                        transaction = Transaction(db=self.db)
                        transaction.set_id(transaction_id[0])
                        self.transactions.append(transaction)


class Product(QObject):
    signal_id_changed = pyqtSignal(int)
    ram_products = dict()

    def __init__(self, db=None):
        self.logger = logging.getLogger(self.__class__.__name__)
        QObject.__init__(self)
        if db:
            self.db = db
        else:
            self.logger.warning(f'connecting to the database using NEW HANDLER')
            self.db = ConnectionDB()

        # SQL TABLE
        self.id = None
        self.product = None
        # LISTS
        self.ids = []
        self.products = []
        self.transactions = []

    def reset(self):
        self.id = None
        self.product = None
        self.loaded = False

    def set_id(self, id):
        self.reset()
        self.logger.debug(f'setting id. was {self.id} is {id}')
        self.id = id
        self.load_data()
        self.signal_id_changed.emit(id)

    def set_id_by_index(self, index):
        self.set_id(self.ids[index])

    def load_products(self):
        self.logger.debug(f'loading products')
        self.ids.clear()
        self.products.clear()
        self.ids.append(None)
        self.products.append(" ")
        cmd = 'SELECT id, product FROM products'
        param = []
        try:
            self.db.execute(cmd, param)
        except Exception as err:
            self.logger.exception(f'FAILED loading products.\nError occured')
        else:
            for element in self.db.fetchall():
                self.ids.append(element[0])
                self.products.append(element[1])

    def load_products_for_customer_only(self, customer_id):
        self.ids.clear()
        self.products.clear()
        self.ids.append(None)
        self.products.append(" ")
        if customer_id is None:
            self.logger.debug(f"Can't load products for customer only while customer id is {customer_id}")
        else:
            cmd = '''SELECT 
                        product_id 
                    FROM 
                        customer_product 
                    WHERE 
                        customer_id=?'''
            param =[customer_id, ]
            try:
                self.db.execute(cmd, param)
            except Exception as err:
                self.logger.exception(f'FAILED loading products for customer only.\nError occured')
            else:
                customer_products = self.db.fetchall()
                for product in customer_products:
                    cmd = '''SELECT
                                id,
                                product
                            FROM
                                products
                            WHERE
                                id=?'''
                    param =[product[0], ]
                    try:
                        self.db.execute(cmd, param)
                    except Exception as err:
                        self.logger.exception(f'FAILED loading product for customer only (inner).\nError occured')
                    else:
                        for element in self.db.fetchall():
                            self.ids.append(element[0])
                            self.products.append(element[1])

    def load_data(self):
        if self.id in self.ram_products.keys():
            print('(PRODUCT) loading data from RAM')
            self.product = self.ram_products[self.id].product
        else:
            ram_product = Product(db=self.db)
            if self.id is not None:
                print('(PRODUCT) loading data from DATABASE')
                self.product = None
                cmd = '''SELECT
                            product
                        FROM
                            products
                        WHERE
                            id = ?'''
                param = [self.id, ]
                try:
                    self.db.execute(cmd, param)
                except Exception:
                    self.logger.exception(f'Error while trying to execute command on database')
                else:
                    try:
                        sql_res = self.db.fetchone()
                    except Exception:
                        self.logger.exception(f'Error while trying to fetching data from database')
                    else:
                        self.product = sql_res[0]
                        ram_product.product = self.product

                        self.ram_products[self.id] = ram_product

    def load_transactions(self, filter_show_released=False):
        self.transactions = []
        cmd = '''SELECT
                    id
                FROM
                    transactions
                WHERE
                    product_id=?'''
        param = [self.id, ]

        if filter_show_released is True:
            cmd += 'AND (released=1 OR released=0)'
        else:
            cmd += 'AND released=0'

        try:
            self.db.execute(cmd, param)
        except Exception:
            self.logger.exception(f'Exception thrown while trying to execute command on database')
        else:
            try:
                sql_res = self.db.fetchall()
            except Exception:
                self.logger.exception(f'Exception while trying to download data from database')
            else:
                for element in sql_res:
                    transaction = Transaction(db=self.db)
                    transaction.set_id(element[0])
                    transaction.load_data_basic()
                    self.transactions.append(transaction)


class Transaction_type(QObject):
    signal_id_changed = pyqtSignal(int)
    ram_transaction_types = dict()

    def __init__(self, db=None):
        self.logger = logging.getLogger(self.__class__.__name__)
        QObject.__init__(self)

        if db:
            self.db = db
        else:
            self.logger.warning(f'connecting to the database using NEW HANDLER')
            self.db = ConnectionDB()

        # SQL TABLE
        self.id = None
        self.type = None
        # LISTS
        self.ids = []
        self.types = []

    def reset(self):
        self.logger.debug(f'reseting atributes')
        self.id = None
        self.type = None
        self.loaded = False

    def set_id(self, id):
        self.reset()
        self.id = id
        self.load_data()
        self.signal_id_changed.emit(self.id)

    def set_id_by_index(self, index):
        self.set_id(self.ids[index])

    def load_types(self):
        self.ids.clear()
        self.types.clear()
        self.ids.append(None)
        self.types.append(" ")
        cmd = 'SELECT id, type FROM trans_type'
        param = []
        try:
            self.db.execute(cmd, param)
        except Exception:
            self.logger.exception(f'FAILED loading types.\nError occured')
        else:
            for element in self.db.fetchall():
                self.ids.append(element[0])
                self.types.append(element[1])

    def load_data(self):
        if self.id in self.ram_transaction_types.keys():
            print('(TRANSACTION TYPE) loading data from RAM')
            self.type = self.ram_transaction_types[self.id].type
        else:
            ram_transaction_type = Transaction_type(db=self.db)
            if self.id is not None:
                print('(TRANSACTION TYPE) loading data from DATABASE')
                cmd = '''SELECT
                            type
                        FROM
                            trans_type
                        WHERE
                            id = ?'''
                param = [self.id, ]
                try:
                    self.db.execute(cmd, param)
                except Exception:
                    self.logger.exception(f'Error while trying execute command on database')
                else:
                    try:
                        sql_res = self.db.fetchone()
                    except Exception:
                        self.logger.exception(f'Error while tryint fetching data from database')
                    else:
                        self.type = sql_res[0]
                        ram_transaction_type.type = self.type

                        self.ram_transaction_types[self.id] = ram_transaction_type


class Discipline(QObject):
    changed_id = pyqtSignal(int, name='discipline_id_changed')
    names_loaded = pyqtSignal()
    ram_disciplines = dict()

    def __init__(self, db=None):
        self.logger = logging.getLogger(self.__class__.__name__)
        QObject.__init__(self)
        if db:
            self.db = db
        else:
            self.logger.warning(f'connecting to the database using NEW HANDLER')
            self.db = ConnectionDB()

        # SQL TABLE
        self.id = None
        self.name = None
        self.customer = Customer(db=db)
        # LISTS
        self.ids = []
        self.names = []

    def __copy__(self):
        cls_copy = Discipline(db=self.db)
        cls_copy.id = self.id
        cls_copy.name = self.name
        return cls_copy

    def reset(self):
        self.id = None
        self.name = None
        self.customer.reset()

    def set_id(self, id):
        self.reset()
        self.id = id
        self.load_data()
        self.changed_id.emit(id)

    def set_id_by_index(self, index):
        self.set_id(self.ids[index])

    def load_names(self):
        self.ids.clear()
        self.names.clear()
        self.ids.append(None)
        self.names.append(" ")
        if self.customer.id:
            cmd = '''SELECT
                        id, 
                        name 
                    FROM 
                        disciplines 
                    WHERE 
                        customer_id = ?'''
            param = [self.customer.id, ]
            try:
                self.db.execute(cmd, param)
            except Exception as err:
                self.logger.exception(f'FAILED loading names.\nError occured')
            else:
                for element in self.db.fetchall():
                    self.ids.append(element[0])
                    self.names.append(element[1])
                self.names_loaded.emit()
        else:
            self.logger.debug(f"Can't load names for disciplines while customer.id is not set")

    def load_data(self):
        if self.id in self.ram_disciplines.keys():
            print('(DISCIPLINE) loading data from RAM')
            self.name = self.ram_disciplines[self.id].name
        else:
            discipline = Discipline(db=self.db)
            if self.id is not None:
                print('(DISCIPLINE) loading data from DATABASE')
                self.name = None
                cmd = '''SELECT
                            name
                        FROM
                            disciplines
                        WHERE
                            id = ?'''
                param = [self.id, ]
                try:
                    self.db.execute(cmd, param)
                except Exception:
                    self.logger.exception(f'Error while trying execute command on database')
                else:
                    try:
                        sql_res = self.db.fetchone()
                    except Exception:
                        self.logger.exception(f'Error while trying fetching data from database')
                    else:
                        self.name = sql_res[0]

                        discipline.name = self.name
                        self.ram_disciplines[self.id] = discipline


class Activity(QObject):
    changed_id = pyqtSignal(int)
    ram_activities = dict()

    def __init__(self, db=None):
        self.logger = logging.getLogger(self.__class__.__name__)
        QObject.__init__(self)
        if db:
            self.db = db
        else:
            self.logger.warning(f'connecting to the database using NEW HANDLER')
            self.db = ConnectionDB()

        # SQL TABLE
        self.id = None
        self.type = None
        self.customer = Customer(db=db)
        self.product = Product(db=db)
        self.transaction_type = Transaction_type(db=db)
        self.discipline = Discipline(db=db)
        # LISTS
        self.ids = []
        self.types = []

    def __copy__(self):
        cls_copy = Activity(db=self.db)
        cls_copy.id = self.id
        cls_copy.type = self.type
        return cls_copy

    def reset(self):
        self.id = None
        self.type = None
        self.customer.reset()
        self.product.reset()
        self.transaction_type.reset()
        self.discipline.reset()

    def set_id(self, id):
        self.reset()
        self.logger.debug(f'setting id. was {self.id} is {id}')
        self.id = id
        self.load_data()
        self.changed_id.emit(id)

    def set_id_by_index(self, index):
        self.set_id(self.ids[index])

    def load_types(self):
        self.ids.clear()
        self.types.clear()
        # self.types = []
        self.ids.append(None)
        self.types.append(" ")
        if self.customer.id and self.product.id and self.transaction_type.id and self.discipline.id:
            cmd = '''SELECT 
                        id, 
                        type 
                    FROM 
                        activities 
                    WHERE 
                        customer_id = ? 
                    AND 
                        product_id = ? 
                    AND 
                        trans_type_id = ? 
                    AND 
                        discipline_id = ?'''
            param = (self.customer.id, self.product.id, self.transaction_type.id, self.discipline.id)
            try:
                self.db.execute(cmd, param)
            except Exception as err:
                self.logger.exception(f'FAILED loading types.\nError occured')
            else:
                for element in self.db.fetchall():
                    self.ids.append(element[0])
                    self.types.append(element[1])

    def load_data(self):
        if self.id in self.ram_activities.keys():
            print("(ACTIVITY) loading data from RAM")
            self.type = self.ram_activities[self.id].type
            self.customer.id = self.ram_activities[self.id].customer.id
            self.product.id = self.ram_activities[self.id].product.id
            self.transaction_type.id = self.ram_activities[self.id].transaction_type.id
            self.discipline.id = self.ram_activities[self.id].discipline.id
        else:
            ram_activity = Activity(db=self.db)
            if self.id is not None:
                print("(ACTIVITY) loading data DATABASE")
                cmd = '''SELECT
                            type,
                            customer_id,
                            product_id,
                            trans_type_id,
                            discipline_id
                        FROM
                            activities
                        WHERE
                            id = ?'''
                param = [self.id, ]
                try:
                    self.db.execute(cmd, param)
                except Exception:
                    self.logger.exception(f'Error while trying to execute command on database')
                else:
                    try:
                        sql_res = self.db.fetchone()
                    except Exception:
                        self.logger.exception(f"Error while tryint to fetching data from database")
                    else:
                        self.type = sql_res[0]
                        self.customer.id = sql_res[1]
                        self.product.id = sql_res[2]
                        self.transaction_type.id = sql_res[3]
                        self.discipline.id = sql_res[4]

                        ram_activity.type = self.type
                        ram_activity.customer.id = self.customer.id
                        ram_activity.product.id = self.product.id
                        ram_activity.transaction_type.id = self.transaction_type.id
                        ram_activity.discipline.id = self.discipline.id

                        self.ram_activities[self.id] = ram_activity


class Transaction(QObject, DateManager):
    signal_locked = pyqtSignal(bool)
    signal_unlocked = pyqtSignal(bool)
    signal_created = pyqtSignal(bool)
    signal_updated = pyqtSignal(bool)
    signal_deleted = pyqtSignal(bool)
    signal_id_changed = pyqtSignal(int)
    signal_status_changed = pyqtSignal(int)
    signal_name_changed = pyqtSignal(str)
    signal_deliverable_changed = pyqtSignal(bool)
    signal_planned_start_changed = pyqtSignal(object)
    signal_planned_end_changed = pyqtSignal(object)
    signal_planned_delivery_changed = pyqtSignal(object)
    signal_true_start_changed = pyqtSignal(object)
    signal_true_end_changed = pyqtSignal(object)
    signal_true_delivery_changed = pyqtSignal(object)
    signal_charge_code_changed = pyqtSignal(str)
    signal_released = pyqtSignal(int)

    names_loaded = pyqtSignal()
    changed_data = pyqtSignal(int)
    signal_message = pyqtSignal(str)

    def __init__(self, db=None):
        self.logger = logging.getLogger(self.__class__.__name__)
        QObject.__init__(self)
        DateManager.__init__(self)

        if db:
            self.db = db
        else:
            self.logger.warning(f'connecting to the database using NEW HANDLER')
            self.db = ConnectionDB()

        # SQL TABLE
        self.id = None
        self.lock = None
        self.locking = Employee(db=db)
        self.name = None
        self.customer = Customer(db=db)
        self.product = Product(db=db)
        self.transaction_type = Transaction_type(db=db)
        self.responsible = Employee(db=db)
        self.planned_start = None
        self.planned_end = None
        self.planned_delivery = None
        self.true_start = None
        self.true_end = None
        self.true_delivery = None
        self.charge_code = None
        self.deliverable = None
        self.released = 0
        self.admins = []

        # LISTS
        self.ids = []
        self.names = []
        self.tasks = []
        self.customers = []
        self.products = []
        self.transaciton_types = []
        self.statuses = []


        # ADDITIONAL
        self.status = None
        self.loaded_basic = False
        self.loaded_extended = False
        self.statistics = {"tasks_count": 0,
                           "delivered_on_time": 0,
                           "delivered_late": 0,
                           "finished_on_time": 0,
                           "finished_late": 0,
                           "ongoing_on_time": 0,
                           "ongoing_late": 0,
                           "planned_on_time": 0,
                           "planned_late": 0,
                           "started_on_time": 0,
                           "started_late": 0,
                           "planning_on_time": 0,
                           "planning_late": 0}

        #FILTERS
        self.filter_show_released_task = False
        self.filter_show_released_checks = False

    def _update_status(self):
        if self.true_delivery is not None:
            self._set_status(status.code["DELIVERED"])
            return

        if self.true_end is not None:
            self._set_status(status.code["FINISHED"])
            return

        # STATUS ONGOING and LATE
        if self.true_start is not None:
            if self.planned_end < date.today():
                self._set_status(status.code["LATE"])
                return
            else:
                self._set_status(status.code["ONGOING"])
                return

        # STATUS SCHEDULED
        if self.planned_start is not None and self.planned_end is not None:
            self._set_status(status.code["SCHEDULED"])
            return

        self._set_status(status.code["ERROR"])

    def _set_status(self, code):
        self.status = code
        self.signal_status_changed.emit(self.status)

    def reset(self):
        self.id = None
        self.lock = None
        self.locking.reset()
        self.name = None
        self.customer.reset()
        self.product.reset()
        self.transaction_type.reset()
        self.responsible.reset()
        self.planned_start = None
        self.planned_end = None
        self.planned_delivery = None
        self.true_start = None
        self.true_end = None
        self.true_delivery = None
        self.charge_code = None
        self.deliverable = None
        self.released = 0
        self.admins.clear()

        self.names = []
        self.tasks = []
        self.customers = []
        self.products = []
        self.transaciton_types = []
        self.statuses = []

        # ADDITIONAL
        self.status = None
        self.loaded_basic = False
        self.loaded_extended = False

        self.statistics = self.statistics.fromkeys(self.statistics, 0)

    def add_admin(self, admin):
        if admin.login not in self.admins:
            self.admins.append(admin.login)

    def remove_admin(self, admin):
        if admin.login in self.admins:
            self.admins.remove(admin.login)

    def get_admins_logins(self):
        logins = ""
        for i, admin_login in enumerate(self.admins):
            if i != len(self.admins)-1:
                logins += admin_login + ";"
            else:
                logins += admin_login

        if logins != "":
            return logins
        else:
            return None

    def set_lock(self):
        if (self.lock == 1 and self.locking.login == os.getlogin()) or self.lock == 0:
            cmd = '''UPDATE
                        transactions
                    SET
                        lock = 1,
                        locking_login = ?
                    WHERE
                        id = ?
                    '''
            param = [os.getlogin().upper(), self.id]
            try:
                self.db.execute(cmd, param)
                self.db.commit()
            except Exception:
                self.signal_locked.emit(False)
                self.logger.exception("FAILED locking transaction.\nError occured")
            else:
                self.lock = 1
                self.locking.set_login(os.getlogin())
                self.logger.info(f'transaction locked successfully with id#{self.db.cursor.lastrowid}')
                self.signal_message.emit("Transaction locked successfully")
                self.signal_locked.emit(True)
        else:
            self.signal_locked.emit(False)
            self.signal_message.emit(f"This transaction is already locked by {self.locking.full_name}")

    def set_unlock(self):
        if (self.lock == 1 and self.locking.login == os.getlogin()) or self.lock == 0:
            cmd = '''UPDATE
                        transactions
                    SET
                        lock = 0,
                        locking_login = NULL
                    WHERE
                        id = ?
                    '''
            param = [self.id]
            try:
                self.db.execute(cmd, param)
                self.db.commit()
            except Exception:
                self.signal_unlocked.emit(False)
                self.logger.exception("FAILED unlocking transaction.\nError occured")
            else:
                self.lock = None
                self.locking.set_login(None)
                self.logger.info(f'transaction unlocking successfully with id#{self.db.cursor.lastrowid}')
                self.signal_message.emit("Transaction unlocked successfully")
                self.signal_unlocked.emit(True)
        else:
            self.signal_unlocked(False)
            self.signal_message.emit(f"This transaction is locked by {self.locking.full_name} and can't be unlocked by you")

    def set_id(self, id):
        self.reset()
        self.id = id
        self.signal_id_changed.emit(self.id)

    def set_id_by_index(self, index):
        self.set_id(self.ids[index])

    def set_name(self, name_to_set):
        if name_to_set != self.name:
            self.name = str(name_to_set)
            self.signal_name_changed.emit(self.name)

    def set_customer_id(self, id):
        if id != self.customer.id:
            self.customer.id = id

    def set_product_id(self, id):
        if id != self.product.id:
            self.product.id = id

    def set_transaction_type(self, id):
        if id != self.transaction_type.id:
            self.transaction_type = id

    def set_deliverable(self, deliverable):
        if deliverable != self.deliverable:
            self.deliverable = int(deliverable)
            self._update_status()
            self.signal_deliverable_changed.emit(self.deliverable)

    def set_planned_start(self, date_to_set):
        if date_to_set is None:
            self.planned_start = None
            self._update_status()
            self.signal_planned_start_changed.emit(self.planned_start)
        else:
            if isinstance(date_to_set, date):
                if date_to_set != self.planned_start:
                    self.planned_start = date_to_set
                    self._update_status()
                    self.signal_planned_start_changed.emit(self.planned_start)
            else:
                self.logger.exception(f'Error occured.')

    def set_planned_end(self, date_to_set):
        if date_to_set is None:
            self.planned_end = None
            self._update_status()
            self.signal_planned_end_changed.emit(self.planned_end)
        else:
            if isinstance(date_to_set, date):
                if date_to_set != self.planned_end:
                    self.planned_end = date_to_set
                    self._update_status()
                    self.signal_planned_end_changed.emit(self.planned_end)
            else:
                raise ValueError ("date_to_set must be type(Date)")

    def set_planned_delivery(self, date_to_set):
        if date_to_set is None:
            self.planned_delivery = None
            self._update_status()
            self.signal_planned_delivery_changed.emit(self.planned_delivery)
        else:
            if isinstance(date_to_set, date):
                if date_to_set != self.planned_delivery:
                    self.planned_delivery = date_to_set
                    self._update_status()
                    self.signal_planned_delivery_changed.emit(self.planned_delivery)
            else:
                raise ValueError("date_to_set must be type(Date)")

    def set_true_start(self, date_to_set):
        if date_to_set is None:
            self.true_start = None
            self._update_status()
            self.signal_true_start_changed.emit(self.true_start)
        else:
            if isinstance(date_to_set, date):
                if date_to_set != self.true_start:
                    self.true_start = date_to_set
                    self._update_status()
                    self.signal_true_start_changed.emit(self.true_start)
            else:
                raise ValueError("date_to_set must be type(Date)")

    def set_true_end(self, date_to_set):
        if date_to_set is None:
            self.true_end = None
            self._update_status()
            self.released_check()
            self.signal_true_end_changed.emit(self.true_end)
        else:
            if isinstance(date_to_set, date):
                if date_to_set != self.true_end:
                    self.true_end = date_to_set
                    self._update_status()
                    self.signal_true_end_changed.emit(self.true_end)
            else:
                raise ValueError("date_to_set must be type(Date)")

    def set_true_delivery(self, date_to_set):
        if date_to_set is None:
            self.true_delivery = None
            self._update_status()
            self.released_check()
            self.signal_true_delivery_changed.emit(self.true_delivery)
        else:
            if isinstance(date_to_set, date):
                if date_to_set != self.true_delivery:
                    self.true_delivery = date_to_set
                    self._update_status()
                    self.signal_true_delivery_changed.emit(self.true_delivery)
            else:
                raise ValueError("date_to_set must be type(Date)")

    def set_charge_code(self, charge_code):
        if charge_code != self.charge_code:
            self.charge_code = charge_code
            self.signal_charge_code_changed.emit(self.charge_code)

    def set_released(self, released):
        if released != self.released:
            self.released = int(released)
            self.signal_released.emit(self.released)
            self._update_status()

    def released_check(self):
        if self.deliverable and self.true_delivery is not None:
            self.set_released(1)
            return
        if not self.deliverable and self.true_end is not None:
            self.set_released(1)
            return
        self.set_released(0)

    def load_names_by_customer(self, show_active_only=True):
        self.ids.clear()
        self.names.clear()
        self.ids.append(None)
        self.names.append(" ")

        if show_active_only is True:
            cmd = '''SELECT
                        id,
                        name
                    FROM
                        transactions
                    WHERE
                        customer_id = ?
                    AND
                        true_start IS NOT NULL
                    AND
                        true_end IS NULL'''
        else:
            cmd = '''SELECT
                        id,
                        name
                    FROM
                        transactions
                    WHERE
                        customer_id = ?
                        '''
        param = (self.customer.id, )
        try:
            self.db.execute(cmd, param)
        except Exception as err:
            print("Error in CLASS Transaction METHOD load_names_by_customer: " + str(err))
        else:
            for element in self.db.fetchall():
                self.ids.append(element[0])
                self.names.append(element[1])

    def load_names_by_customer_product_type(self):
        self.ids.clear()
        self.names.clear()

        cmd = '''SELECT
                    id,
                    name
                FROM
                    transactions
                WHERE
                    customer_id=?
                AND
                    product_id=?
                AND
                    trans_type_id=?
                AND
                    true_start IS NOT NULL
                AND
                    true_end IS NULL'''
        param = [self.customer.id, self.product.id, self.transaction_type.id]
        try:
            self.db.execute(cmd, param)
        except Exception as err:
            pass #logger here
        else:
            for element in self.db.fetchall():
                self.ids.append(element[0])
                self.names.append(element[1])

    def load_data_basic(self):
        if self.id and not self.loaded_basic:
            print("(TRANSACTION) loading data [id={}]".format(self.id))
            cmd = '''SELECT
                        name,
                        lock,
                        locking_login,
                        customer_id,
                        product_id,
                        trans_type_id,
                        responsible_login,
                        planned_start,
                        planned_end,
                        planned_delivery,
                        true_start,
                        true_end,
                        true_delivery,
                        charge_code,
                        deliverable,
                        released,
                        admins
                    FROM
                        transactions
                    WHERE
                        id = ?'''

            param = [self.id, ]
            try:
                self.db.execute(cmd, param)
            except Exception:
                self.logger.exception("Error while trying to execute command on database")
            else:
                try:
                    sql_res = self.db.fetchone()
                except Exception:
                    self.logger.exception("Error while trying to fetching data from database")
                else:
                    self.name = sql_res[0]
                    self.lock = sql_res[1]
                    self.locking.login = sql_res[2]
                    self.customer.id = sql_res[3]
                    self.customer.load_data()
                    self.product.id = sql_res[4]
                    self.product.load_data()
                    self.transaction_type.id = sql_res[5]
                    self.transaction_type.load_data()
                    self.responsible.login = sql_res[6]
                    self.responsible.load_data()
                    self.planned_start = self.convert_to_date_class(sql_res[7])
                    self.planned_end = self.convert_to_date_class(sql_res[8])
                    self.planned_delivery = self.convert_to_date_class(sql_res[9])
                    self.true_start = self.convert_to_date_class(sql_res[10])
                    self.true_end = self.convert_to_date_class(sql_res[11])
                    self.true_delivery = self.convert_to_date_class(sql_res[12])
                    self.charge_code = sql_res[13]
                    self.deliverable = int(sql_res[14])
                    self.released = int(sql_res[15])
                    try:
                        self.admins = sql_res[16].split(';')
                    except AttributeError:
                        pass

                    self._update_status()
                    self.loaded_basic = True

    def load_data_childs(self, employee_login=None):
        print("(TRANSACTION) loading data - extended [id={}]".format(self.id))
        if employee_login:
            self._load_tasks_where_employee_is_responsible(employee_login=employee_login)
        else:
            self._load_all_tasks_for_transaction()
        self.loaded_extended = True

    def has_childs(self):
        cmd = '''SELECT
                    id
                FROM
                    tasks
                WHERE
                    transaction_id=?'''

        if self.filter_show_released_task is True:
            cmd += 'AND (released=1 OR released=0)'
        else:
            cmd += 'AND released=0'

        param = [self.id]
        try:
            self.db.execute(cmd, param)
        except Exception as err:
            self.logger.exception(f'FAILED loading tasks for transaction.\nError occured: ')
        else:
            try:
                sql_res = self.db.fetchall()
            except Exception:
                self.logger.exception(f'Error while trying to download data from database')
            else:
                if len(sql_res) == 0:
                    return False
                else:
                    return True

    def _load_all_tasks_for_transaction(self):
        self.tasks = []
        cmd = '''SELECT
                    id
                FROM
                    tasks
                WHERE
                    transaction_id=?'''
        param = [self.id, ]

        if self.filter_show_released_task is True:
            cmd += 'AND (released=1 OR released=0)'
        else:
            cmd += 'AND released=0'

        try:
            self.db.execute(cmd, param)
        except Exception:
            self.logger.exception(f'FAILED loading tasks for transaction.\nError occured: ')
        else:
            try:
                sql_res = self.db.fetchall()
            except Exception:
                self.logger.exception(f'Error while trying to download data from database')
            else:
                for element in sql_res:
                    task = Task(db=self.db)
                    task.set_id(element[0])
                    task.load_data_basic()
                    # task.load_data_extended()
                    self.tasks.append(task)

    def _load_tasks_where_employee_is_responsible(self, employee_login):
        self.tasks = []
        if self.filter_show_released_task:
            filter_task = f' '
        else:
            filter_task = f'AND released = 0'

        if self.filter_show_released_checks:
            filter_check = f' '
        else:
            filter_check = f'AND released = 0'

        cmd = f'SELECT' \
              f' id ' \
              f'FROM' \
              f' tasks ' \
              f'WHERE' \
              f' transaction_id = ? ' \
              f'AND' \
              f' id ' \
              f'IN' \
              f'(SELECT' \
              f' task_id ' \
              f'FROM' \
              f' checks ' \
              f'WHERE' \
              f' responsible_login = ? ' \
              f'{filter_check} )' \
              f'OR' \
              f'( transaction_id = ? ' \
              f'AND' \
              f' responsible_login = ? ' \
              f' {filter_task})'

        param = [self.id, employee_login, self.id, employee_login]

        try:
            self.db.execute(cmd, param)
        except Exception :
            self.logger.exception(f'Error while trying to send command to database')
        else:
            try:
                sql_res = self.db.fetchall()
            except Exception:
                self.logger.exception(f'Error while trying to download data from database')
            else:
                for element in sql_res:
                    task = Task(db=self.db)
                    task.set_id(element[0])
                    task.load_data_basic()
                    self.tasks.append(task)

    def _load_comments(self):
        self.comments = []
        cmd = '''SELECT
                    id
                FROM
                    comments
                WHERE
                    transaction_id=?
                AND
                    task_id IS NULL'''
        param = [self.id]
        try:
            self.db.execute(cmd, param)
        except Exception:
            self.logger.exception(f'FAILED loading comments for task.\nError occured: ')
        else:
            sql_res = self.db.fetchall()
            for element in sql_res:
                comment = Comment(db=self.db)
                comment.set_id(element[0])
                self.comments.append(comment)

    def update_data(self):
        if self.ready_to_save():
            self.released_check()
            cmd = '''UPDATE
                        transactions
                    SET
                        name = ?,
                        customer_id = ?,
                        product_id = ?,
                        trans_type_id = ?,
                        responsible_login = ?,
                        planned_start = ?,
                        planned_end = ?,
                        planned_delivery = ?,
                        true_start = ?,
                        true_end = ?,
                        true_delivery = ?,
                        charge_code = ?,
                        deliverable = ?,
                        released = ?,
                        admins = ?
                    WHERE
                        id = ?
                    '''
            param = [self.name,
                     self.customer.id,
                     self.product.id,
                     self.transaction_type.id,
                     self.responsible.login,
                     self.planned_start,
                     self.planned_end,
                     self.planned_delivery,
                     self.true_start,
                     self.true_end,
                     self.true_delivery,
                     self.charge_code,
                     self.deliverable,
                     self.released,
                     self.get_admins_logins(),
                     self.id]
            try:
                self.db.execute(cmd, param)
                self.db.commit()
            except Exception:
                self.signal_updated.emit(False)
                self.logger.exception("FAILED updating transaction.\nError occured")
            else:
                self.logger.info(f'transaction updated successfully with id#{self.db.cursor.lastrowid}')
                self.signal_message.emit("Transaction updated successfully")
                self.signal_updated.emit(True)

    def register_data(self):
        if self.ready_to_save():
            cmd = '''INSERT INTO 
                        transactions 
                    VALUES(NULL, 0, NULL, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)'''

            param = [self.name,
                     self.customer.id,
                     self.product.id,
                     self.transaction_type.id,
                     self.responsible.login,
                     self.planned_start,
                     self.planned_end,
                     self.planned_delivery,
                     self.true_start,
                     self.true_end,
                     self.true_delivery,
                     self.charge_code,
                     self.deliverable,
                     self.released,
                     self.get_admins_logins()
                     ]

            try:
                self.db.execute(cmd, param)
                self.db.commit()
            except Exception:
                self.signal_created.emit(False)
                self.logger.exception("FAILED save new transaction.\nError occured")
            else:
                self.logger.info(f'new transaction saved successfully with id#{self.db.cursor.lastrowid}')
                self.signal_message.emit("Transaction saved successfully")
                self.signal_created.emit(True)

    def delete_data(self):
        if self.id is not None:
            cmd = f'DELETE FROM transactions WHERE id=?'
            param = [self.id,]
            try:
                self.db.execute(cmd, param)
                self.db.commit()
            except Exception:
                self.signal_deleted.emit(False)
                self.logger.exception(f'Error while trying to execute delete command on database')
            else:
                self.signal_deleted.emit(True)

    def ready_to_save(self):
        if not self.name:
            self.signal_message.emit("Name can't be empty")
            return False

        if not self.transaction_type.id:
            self.signal_message.emit("Transaction_type.id can't be None")
            return False

        if not self.product.id:
            self.signal_message.emit("Product.id can't be None")
            return False

        if not self.customer.id:
            self.signal_message.emit("Customer.id can't be None")
            return False

        if not self.responsible.login:
            self.signal_message.emit("Responsible person can't be empty")
            return False

        if self.deliverable is None:
            self.signal_message.emit(f'Please select YES/NO for deliverable option')
            return False

        if self.deliverable:
            if not self.planned_delivery:
                self.signal_message.emit("If transaction is Deliverable then Planned Delivery date can't be empty")
                return False

        if not self.planned_start:
            self.signal_message.emit("Planned start date can't be None")
            return False

        if not self.planned_end:
            self.signal_message.emit("Planned end date can't be None")
            return False

        if self.planned_end is not None and self.planned_delivery is not None:
            if (self.planned_delivery - self.planned_end).days < 0:
                self.signal_message.emit("Planned Delivery can't be before Planned End")
                return False

        if self.planned_start is not None and self.planned_end is not None:
            if (self.planned_end - self.planned_start).days < 0:
                self.signal_message.emit("Planned End can't be before Planned Start")
                return False


        # if self.lock == 1 and self.locking.login != os.getlogin(): #TODO: Czasowo wylaczone
        #     self.signal_message.emit(f"Current transaction is locked for editing by {self.locking.full_name}")
        #     return False

        return True

    def calculate_statistics(self):

        self._load_all_tasks_for_transaction()
        self.statistics = self.statistics.fromkeys(self.statistics, 0)

        self.statistics["tasks_count"] = len(self.tasks)

        for task in self.tasks:
            print(f'Calculate stats for task_id:{task.id}')
            if task.status >= status.code["ONGOING"]:
                if task.planned_start >= task.true_start:
                    self.statistics["planning_on_time"] +=1
                else:
                    self.statistics["planning_late"] +=1

            if task.status <= status.code["SCHEDULED"]:
                if task.planned_start <= datetime.now().date():
                    self.statistics["planned_on_time"] += 1
                else:
                    self.statistics["planned_late"] += 1
            elif task.status >= status.code["ONGOING"] and task.status < status.code["FINISHED"]:
                if task.deliverable == 1:
                    if task.planned_delivery >= datetime.now().date():
                        self.statistics["ongoing_on_time"] += 1
                    else:
                        self.statistics["ongoing_late"] += 1
                else:
                    if task.planned_end >= datetime.now().date():
                        self.statistics["ongoing_on_time"] += 1
                    else:
                        self.statistics["ongoing_late"] += 1
            elif task.status >= status.code["FINISHED"] and task.status < status.code["DELIVERED"]:
                if task.deliverable == 1:
                    if task.released == 1:
                        if task.true_delivery <= task.planned_delivery:
                            self.statistics["delivered_on_time"] += 1
                        else:
                            self.statistics["delivered_late"] += 1
                    else:
                        if task.planned_delivery <= datetime.now().date():
                            self.statistics["ongoing_on_time"] += 1
                        else:
                            self.statistics["ongoing_late"] += 1
                else:
                    if task.true_end <= task.planned_end:
                        self.statistics["finished_on_time"] +=1
                    else:
                        self.statistics["finished_late"] += 1
            elif task.status == status.code["DELIVERED"]:
                # if task.deliverable == 1:
                #     if task.released == 1:
                        if task.true_delivery <= task.planned_delivery:
                            self.statistics["delivered_on_time"] += 1
                        else:
                            self.statistics["delivered_late"] += 1
                    # else:
                    #     if task.planned_delivery <= datetime.now().date():
                    #         self.statistics["ongoing_on_time"] += 1
                    #     else:
                    #         self.statistics["ongoing_late"] += 1
                # else:
                #     if task.true_end <= task.planned_end:
                #         self.statistics["finished_on_time"] +=1
                #     else:
                #         self.statistics["finished_late"] += 1


class Employee(QObject, DateManager):
    signal_login_changed = pyqtSignal(str)
    ram_employees = dict()

    def __init__(self, db=None):
        DateManager.__init__(self)
        QObject.__init__(self)
        self.logger = logging.getLogger(self.__class__.__name__)
        if db:
            self.db = db
        else:
            self.logger.warning(f'connecting to the database using NEW HANDLER')
            self.db = ConnectionDB()

        # SQL TABLE
        self.login = None
        self.status_id = None
        self.access_id = None
        self.emp_number = None
        self.first_name = None
        self.last_name = None
        self.e_mail = None
        self.title = None
        self.department_id = None
        self.checker_rights = None
        self.working_hours = None
        self.vacation = None
        self.phone_ext = None
        self.mobile = None

        # LISTS
        self.logins = []
        self.stack_logins = []
        self.full_names = []
        self.stack_tasks = []
        self.other_employees = []
        self.workload = dict()

        # OTHER
        self.full_name = None
        self.stack_task = None

    def __copy__(self):
        cls_copy = Employee(db=self.db)
        cls_copy.login = self.login
        cls_copy.full_name = self.full_name
        return cls_copy

    def set_login_by_index(self, index):
        self.set_login(self.logins[index])

    def set_login(self, login):
        if login is None or login == "":
            self.login = None
        else:
            self.login = str(login).upper()
        self.load_data() #TODO: do usunięcia, trzeba sprawdzić gdzie wywolac ręcznie
        self.signal_login_changed.emit(self.login)

    def reset(self):
        self.logger.debug(f'reset variables')
        self.login = None
        self.status_id = None #TODO: Need class for status
        self.access_id = None #TODO: Need class for access
        self.emp_number = None
        self.first_name = None
        self.last_name = None
        self.e_mail = None
        self.title = None
        self.department_id = None
        self.checker_rights = None
        self.working_hours = None
        self.vacation = None
        self.phone_ext = None
        self.mobile = None

        self.logins = []

    def load_stack_tasks(self):
        self.stack_logins.clear()
        self.stack_tasks.clear()
        self.stack_logins.append(None)
        self.stack_tasks.append(" ")

        cmd = '''SELECT 
                    login, 
                    first_name,
                    last_name
                FROM 
                   employees 
               WHERE 
                   login
                LIKE ?
               '''
        param = ["%STACK_TASK%", ]
        try:
            self.db.execute(cmd, param)
        except Exception:
            self.logger.exception(f'Error while trying to load stack task')
        else:
            sql_res = self.db.fetchall()
            for element in sql_res:
                self.stack_logins.append(element[0])
                self.stack_tasks.append(element[1])

    def load_full_names(self, filter_str="", status=1, stack=True): #TODO: class for employee_status ?
        self.logger.info(f'loading variables: login, first_name, last_name')
        self.logger.debug(f'cleaning lists')
        self.logins.clear()
        self.full_names.clear()
        self.logins.append(None)
        self.full_names.append(" ")

        try:
            if not filter_str:
                cmd = '''SELECT 
                            login, 
                            first_name, 
                            last_name 
                        FROM 
                            employees 
                        WHERE 
                            status_id = ?'''
                param = [status, ]
                try:
                    self.db.execute(cmd, param)
                except:
                    pass
            else:
                cmd = '''SELECT 
                            login, 
                            first_name, 
                            last_name 
                        FROM 
                            employees 
                        WHERE 
                            status_id = ? 
                        AND 
                            last_name LIKE "%{?}%" 
                        OR 
                            first_name LIKE "%{?}%"''' #TODO check if its working. was "%{str}%}
                param = [status, filter_str, ]
                self.db.execute(cmd, param)

            for element in self.db.fetchall():
                if stack is True:
                    self.logins.append(element[0])
                    self.full_names.append(element[2] + " " + element[1])
                else:
                    if not element[0][:10] == "STACK_TASK":
                        self.logins.append(element[0])
                        self.full_names.append(element[2] + " " + element[1])

            # self.logins, self.full_names = (list(t) for t in zip(*sorted(zip(self.logins, self.full_names))))
            self.full_names, self.logins = zip(*sorted(zip(self.full_names, self.logins)))
            self.full_names = list(self.full_names)
            self.logins = list(self.logins)

        except Exception as err:
            pass

    def load_checkers(self):
        self.logins.clear()
        self.full_names.clear()
        self.logins.append(None)
        self.full_names.append(" ")
        cmd = '''SELECT 
                    login, 
                    first_name, 
                    last_name 
                FROM 
                    employees 
                WHERE 
                    checker_rights = ?'''
        param = [1, ]
        try:
            self.db.execute(cmd, param)
        except Exception as err:
            self.err_log.write(str(err))
        else:
            for element in self.db.fetchall():
                self.logins.append(element[0])
                self.full_names.append(element[2] + " " + element[1])

            self.full_names, self.logins = (list(t) for t in zip(*sorted(zip(self.full_names, self.logins))))

    def load_data(self):
        if self.login in self.ram_employees.keys():
            print('(EMPLOYEE) loading from RAM')
            emp_data = self.ram_employees[self.login]
            self.login = emp_data.login
            self.status_id = emp_data.status_id
            self.access_id = emp_data.access_id
            self.emp_number = emp_data.emp_number
            self.first_name = emp_data.first_name
            self.last_name = emp_data.last_name
            self.e_mail = emp_data.e_mail
            self.title = emp_data.title
            self.department_id = emp_data.department_id
            self.checker_rights = emp_data.checker_rights
            self.working_hours = emp_data.working_hours
            self.phone_ext = emp_data.phone_ext
            self.mobile = emp_data.mobile
            self.vacation = emp_data.vacation
            self.full_name = emp_data.full_name
        else:
            emp_data = Employee(db=self.db)
            if self.login != None:
                print('(EMPLOYEE) loading data from DATABASE')
                cmd = '''SELECT
                            login,
                            status_id,
                            access_id,
                            emp_number,
                            first_name,
                            last_name,
                            e_mail,
                            title,
                            dep_id,
                            checker_rights,
                            part_time,
                            desk_phone_ext,
                            mobile_phone,
                            vacation_aval_days                        
                        FROM
                            employees
                        WHERE
                            login = ?'''
                param = [self.login, ]
                try:
                    self.db.execute(cmd, param)
                except Exception as err:
                    print("Error in CLASS Employee METHOD load_info: " + str(err))
                else:
                    sql_res = self.db.fetchone()
                    if sql_res is not None:
                        self.login = sql_res[0]
                        self.status_id = sql_res[1]
                        self.access_id = sql_res[2]
                        self.emp_number = sql_res[3]
                        self.first_name = sql_res[4]
                        self.last_name = sql_res[5]
                        self.e_mail = sql_res[6]
                        self.title = sql_res[7]
                        self.department_id = sql_res[8]
                        self.checker_rights = sql_res[9]
                        self.working_hours = sql_res[10]
                        self.phone_ext = sql_res[11]
                        self.mobile = sql_res[12]
                        self.vacation = sql_res[13]

                        self.full_name = self.first_name + " " + self.last_name

                emp_data.login = self.login
                emp_data.status_id = self.status_id
                emp_data.access_id = self.access_id
                emp_data.emp_number = self.emp_number
                emp_data.first_name = self.first_name
                emp_data.last_name = self.last_name
                emp_data.e_mail = self.e_mail
                emp_data.title = self.title
                emp_data.department_id = self.department_id
                emp_data.checker_rights = self.checker_rights
                emp_data.working_hours = self.working_hours
                emp_data.phone_ext = self.phone_ext
                emp_data.mobile = self.mobile
                emp_data.vacation = self.vacation
                emp_data.full_name = self.full_name

                self.ram_employees[self.login] = emp_data

            else:
                self.first_name = None
                self.last_name = None
                self.full_name = None

    def load_workload(self):
        self.workload.clear()
        if self.login:
            cmd = '''SELECT
                        planned_start,
                        planned_end,
                        hours_planned
                    FROM
                        tasks
                    WHERE
                        responsible_login = ?'''
            param = [self.login, ]

            try:
                self.db.execute(cmd, param)
            except Exception:
                self.logger.exception(f"Error")
            else:
                for element in self.db.fetchall():
                    planned_start_date = self.convert_to_date_class(element[0])
                    planned_end_date = self.convert_to_date_class(element[1])
                    hours_planned = element[2]

                    daydiff = planned_end_date.weekday() - planned_start_date.weekday()
                    working_task_days = int(
                        ((planned_end_date - planned_start_date).days - daydiff) / 7 * 5 + min(daydiff, 5) - (
                            max(planned_end_date.weekday() - 4, 0) % 5)) + 1
                    task_workload = (hours_planned / 8 / working_task_days) * 100


                    for single_date in self.daterange(planned_start_date, planned_end_date):
                        if single_date in self.workload:
                            self.workload[single_date] += task_workload
                        else:
                            self.workload[single_date] = task_workload

            cmd = '''SELECT
                        planned_start,
                        planned_end,
                        hours_planned
                    FROM
                        checks
                    WHERE
                        responsible_login = ?'''
            param = [self.login, ]

            try:
                self.db.execute(cmd, param)
            except Exception:
                self.logger.exception(f"Error")
            else:
                for element in self.db.fetchall():
                    planned_start_date = self.convert_to_date_class(element[0])
                    planned_end_date = self.convert_to_date_class(element[1])
                    hours_planned = element[2]
                    if planned_start_date is not None and planned_end_date is not None:
                        daydiff = planned_end_date.weekday() - planned_start_date.weekday()
                        working_task_days = int(
                            ((planned_end_date - planned_start_date).days - daydiff) / 7 * 5 + min(daydiff, 5) - (
                                max(planned_end_date.weekday() - 4, 0) % 5)) + 1
                        try:
                            task_workload = (hours_planned / 8 / working_task_days) * 100
                        except Exception as err:
                            pass

                        for single_date in self.daterange(planned_start_date, planned_end_date):
                            if single_date in self.workload:
                                self.workload[single_date] += task_workload
                            else:
                                self.workload[single_date] = task_workload

    def load_department_employees_workload(self):
        self.other_employees.clear()
        if self.has_account:
            cmd = '''SELECT
                        login
                    FROM
                        employees
                    WHERE
                        dep_id = ?'''
            param = [self.department_id, ]
            try:
                self.db.execute(cmd, param)
            except Exception as err:
                self.err_log.write(str(err))
            else:
                sql_res = self.db.fetchall()
                for emp_login in sql_res:
                    employee = Employee(db=self.db)
                    employee.set_login(emp_login[0])
                    employee.load_data()
                    employee.load_workload()
                    self.other_employees.append(employee)

    def has_account(self):
        if self.login:
            cmd = '''SELECT EXISTS(SELECT 1 FROM
                        employees
                    WHERE
                        login = ?)'''
            param = (self.login,)
            try:
                self.db.execute(cmd, param)
            except Exception as err:
                print("Error in CLASS Employee METHOD valid: " + str(err))
            else:
                return self.db.fetchone()[0]


class Task(QObject, DateManager):
    signal_locked = pyqtSignal(bool)
    signal_unlocked = pyqtSignal(bool)
    signal_created = pyqtSignal(bool)
    signal_updated = pyqtSignal(bool)
    signal_deleted = pyqtSignal(bool)
    signal_id_changed = pyqtSignal(int)
    signal_status_changed = pyqtSignal(int)
    signal_name_changed = pyqtSignal(str)
    signal_deliverable_changed = pyqtSignal(bool)
    signal_checker_required_changed = pyqtSignal(bool)
    signal_planned_start_changed = pyqtSignal(object)
    signal_planned_end_changed = pyqtSignal(object)
    signal_planned_delivery_changed = pyqtSignal(object)
    signal_true_start_changed = pyqtSignal(object)
    signal_true_end_changed = pyqtSignal(object)
    signal_true_delivery_changed = pyqtSignal(object)
    signal_hours_planned_changed = pyqtSignal(float)
    signal_hours_utilized_changed = pyqtSignal(float)
    signal_released_changed = pyqtSignal(int)
    signal_message = pyqtSignal(str)

    ram_tasks = dict()

    def __init__(self, db=None):
        self.logger = logging.getLogger(self.__class__.__name__)
        QObject.__init__(self)
        DateManager.__init__(self)
        if db:
            self.db = db
        else:
            self.logger.warning(f'used NEW HANDLER for connection')
            self.db = ConnectionDB()

        #SQL TABLE
        self.id = None
        self.name = None
        self.lock = 0
        self.locking = Employee(db=self.db) #TODO: moze nie trzeba tutaj klasy tylko login
        self.transaction_id = None
        self.discipline = Discipline(db=self.db)
        self.activity = Activity(db=self.db)
        self.creator = Employee(db=self.db) #TODO: moze nie trzeba tutaj klasy tylko login
        self.responsible = Employee(db=self.db) #TODO: moze nie trzeba tutaj klasy tylko login
        self.planned_start = None
        self.planned_end = None
        self.planned_delivery = None
        self.true_start = None
        self.true_end = None
        self.true_delivery = None
        self.hours_planned = 0
        self.hours_utilized = 0
        self.deliverable = None
        self.checker_required = None
        self.released = 0
        self.editable = 0

        #LISTS
        self.ids = []
        self.names = []
        self.statuses = []

        # OTHERS
        self.transaction = Transaction(db=self.db)
        self.status = 0
        self.loaded_basic = False
        self.loaded_extended = False
        self.rework = Rework(db=self.db)
        self.reworks = []
        self.check = Check(db=self.db)
        self.checks = []
        self.comments = []

    def __copy__(self):
        cls_copy = Task(db=self.db)
        cls_copy.id = self.id
        cls_copy.name = self.name
        cls_copy.transaction_id = self.transaction_id
        cls_copy.discipline = copy.copy(self.discipline)
        cls_copy.activity = copy.copy(self.activity)
        cls_copy.responsible = copy.copy(self.responsible)
        cls_copy.planned_start = self.planned_start
        cls_copy.planned_end = self.planned_end
        cls_copy.planned_delivery = self.planned_delivery
        cls_copy.true_start = self.true_start
        cls_copy.true_end = self.true_end
        cls_copy.true_delivery = self.true_delivery
        cls_copy.hours_planned = self.hours_planned
        cls_copy.hours_utilized = self.hours_utilized
        cls_copy.deliverable = self.deliverable
        cls_copy.checker_required = self.checker_required
        cls_copy.released = self.released
        return cls_copy

    def _update_status(self):

        #STATUS DELIVERED
        if self.true_delivery is not None:
            self._set_status(status.code["DELIVERED"])
            return


        #STATUS READY FOR DELIVERY or FINISHED
        if self.checker_required == 0 and self.true_end is not None:

            if self.true_delivery is not None:
                self._set_status(status.code["DELIVERED"])
                return

            if self.true_delivery is None and self.deliverable == 1:
                self._set_status(status.code["READY TO DELIVERY"])
                return

            self._set_status(status.code["FINISHED"])
            return

        #STATUS CHECKING
        if self.checker_required == 1 and self.true_end is not None:

            if self.released == 1:
                self._set_status(status.code["FINISHED"])
                return

            if self.rework.status is not None:
                if self.rework.status > status.code["ASSIGNED"] and self.rework.status < status.code["FINISHED"]:
                    self._set_status(status.code["IN REWORK"])
                    return

            if self.check.status is not None:
                if self.check.status == status.code["RELEASED"]:
                    self._set_status(status.code["READY TO DELIVERY"])
                    return

                if self.check.status == status.code["RETURNED"]:
                    self._set_status(status.code["REWORK NEEDED"])
                    return

                if self.check.status > status.code["READY TO POP"] and self.check.status < status.code["FINISHED"]:
                    self._set_status(status.code["IN CHECK"])
                    return

            if self.check.status == 0:
                self._set_status(status.code["READY TO CHECK"])
                return


        #STATUS ONGOING and LATE
        if self.true_start is not None:
            if self.planned_end < date.today():
                self._set_status(status.code["LATE"])
                return
            else:
                self._set_status(status.code["ONGOING"])
                return

        #STATUS SCHEDULED
        if self.planned_start is not None and self.planned_end is not None:
            self._set_status(status.code["SCHEDULED"])
            return

        #STATUS READY TO POP and ASSIGNED
        if self.responsible.login is not None:
            if self.responsible.login[:10] == "STACK_TASK":
                self._set_status(status.code["READY TO POP"])
                return
            else:
                self._set_status(status.code["ASSIGNED"])
                return

        self._set_status(status.code["ERROR"])

    def _set_status(self, code):
        self.status = code
        self.signal_status_changed.emit(self.status)

    def reset(self):
        self.id = None
        self.name = None
        self.lock = 0
        self.locking.reset()
        self.transaction_id = None
        self.discipline.reset()
        self.activity.reset()
        self.creator.reset()
        self.responsible.reset()
        self.planned_start = None
        self.planned_end = None
        self.planned_delivery = None
        self.true_start = None
        self.true_end = None
        self.true_delivery = None
        self.hours_planned = 0
        self.hours_utilized = 0
        self.deliverable = None
        self.checker_required = None
        self.released = 0
        self.editable = 0
        self.status = 0

        self.transaction.reset()
        self.loaded_basic = False
        self.loaded_extended = False
        self.check.reset()
        self.rework.reset()
        self.reworks = []
        self.checks = []
        self.comments = []

    def set_lock(self):
        if (self.lock == 1 and self.locking.login == os.getlogin().upper()) or self.lock == 0:
            cmd = '''UPDATE
                        tasks
                    SET
                        lock = 1,
                        locking_login = ?
                    WHERE
                        id = ?
                    '''
            param = [os.getlogin().upper(), self.id]
            try:
                self.db.execute(cmd, param)
                self.db.commit()
            except Exception:
                self.signal_locked.emit(False)
                self.logger.exception("FAILED locking task.\nError occured")
            else:
                self.lock = 1
                self.locking.set_login(os.getlogin().upper())
                self.signal_message.emit("Task locked successfully")
                self.signal_locked.emit(True)
        else:
            self.signal_locked.emit(False)
            self.signal_message.emit(f"This task is locked by {self.locking.full_name}")

    def set_unlock(self):
        if (self.lock == 1 and self.locking.login == os.getlogin().upper()) or self.lock == 0:
            cmd = '''UPDATE
                        tasks
                    SET
                        lock = 0,
                        locking_login = NULL
                    WHERE
                        id = ?
                    '''
            param = [self.id]
            try:
                self.db.execute(cmd, param)
                self.db.commit()
            except Exception:
                self.signal_unlocked.emit(False)
                self.logger.exception("FAILED unlocking task.\nError occured")
            else:
                self.lock = 0
                self.locking.set_login(None)
                self.signal_message.emit("Task unlocked successfully")
                self.signal_unlocked.emit(True)
        else:
            self.signal_unlocked.emit(False)
            self.signal_message.emit(f"This task is locked by {self.locking.full_name} and can't be unlocked by you")

    def set_id(self, id):
        self.reset()
        self.id = id
        self.signal_id_changed.emit(self.id)

    def set_id_by_index(self, index):
        self.set_id(self.ids[index])

    def set_name(self, name):
            self.name = name
            self.signal_name_changed.emit(self.name)

    def set_responsible_login(self, login):
        if login != self.responsible.login:
            self.responsible.set_login(login)
            if login[:5] == "STACK":
                self.editable = 1
            else:
                self.editable = 0

    def set_transaction_id(self, id):
        if id != self.transaction_id:
            self.transaction_id = id

    def set_discipline_id(self, id):
        if id != self.discipline_id:
            self.discipline_id = id

    def set_activity_id(self, id):
        if id != self.activity_id:
            self.activity_id = id

    def set_checker_required(self, state):
        if state != self.checker_required:
            self.checker_required = int(state)
            if self.deliverable:
                if self.true_delivery is not None:
                    self.set_released(1)
            else:
                if self.true_end is not None:
                    self.set_released(1)

            self._update_status()
            self.signal_checker_required_changed.emit(self.checker_required)

    def set_deliverable(self, deliverable):
        if deliverable != self.deliverable:
            self.deliverable = int(deliverable)
            self.signal_deliverable_changed.emit(self.deliverable)

    def set_planned_start(self, date_to_set):
        if date_to_set is None:
            self.planned_start = None
            self._update_status()
            self.signal_planned_start_changed.emit(None)
        else:
            if isinstance(date_to_set, date):
                if date_to_set != self.planned_start:
                    self.planned_start = date_to_set
                    self._update_status()
                    self.signal_planned_start_changed.emit(self.planned_start)
            else:
                raise TypeError("must be date not {}".format(type(date_to_set)))

    def set_planned_end(self, date_to_set):
        if date_to_set is None:
            self.planned_end = None
            self._update_status()
            self.signal_planned_end_changed.emit(None)
        else:
            if isinstance(date_to_set, date):
                if date_to_set != self.planned_end:
                    self.planned_end = date_to_set
                    self._update_status()
                    self.signal_planned_end_changed.emit(self.planned_end)
            else:
                raise TypeError("must be date not {}".format(type(date_to_set)))

    def set_planned_delivery(self, date_to_set):
        if date_to_set is None:
            self.planned_delivery = None
            self._update_status()
            self.signal_planned_delivery_changed.emit(self.planned_delivery)
        else:
            if isinstance(date_to_set, date):
                if date_to_set != self.planned_delivery:
                    self.planned_delivery = date_to_set
                    self._update_status()
                    self.signal_planned_delivery_changed.emit(self.planned_delivery)
            else:
                raise TypeError("must be date not {}".format(type(date_to_set)))

    def set_true_start(self, date_to_set):
        if date_to_set is None:
            if self.true_start is not None:
                self.true_start = None
                self._update_status()
                self.signal_true_start_changed.emit(self.true_start)
        else:
            if isinstance(date_to_set, date):
                if date_to_set != self.true_start:
                    self.true_start = date_to_set
                    self._update_status()
                    self.signal_true_start_changed.emit(self.true_start)
            else:
                raise TypeError("must be date not {}".format(type(date_to_set)))

    def set_true_end(self, date_to_set):
        if date_to_set is None:
            if self.true_end is not None:
                self.true_end = None
                self._update_status()
                self.signal_true_end_changed.emit(self.true_end)
        else:
            if isinstance(date_to_set, date):
                if date_to_set != self.true_end:
                    self.true_end = date_to_set
                    self._update_status()
                    self.signal_true_end_changed.emit(self.true_end)
            else:
                raise TypeError("must be date not {}".format(type(date_to_set)))

    def set_true_delivery(self, date_to_set):
        if date_to_set is None:
            if self.true_delivery is not None:
                self.true_delivery = None
                self._update_status()
                self.signal_true_delivery_changed.emit(self.true_delivery)
        else:
            if isinstance(date_to_set, date):
                if date_to_set != self.true_delivery:
                    self.true_delivery = date_to_set
                    self._update_status()
                    self.signal_true_delivery_changed.emit(self.true_delivery)
            else:
                raise TypeError("must be date not {}".format(type(date_to_set)))

    def set_avability(self, avability):
        if avability != self.avability:
            self.avability = round(avability, 2) #TODO: czy to jest dobrze liczone ?
            self.signal_avability_changed.emit(self.avability)

    def set_hours_planned(self, hours):
            self.hours_planned = round(hours, 1)
            self.signal_hours_planned_changed.emit(self.hours_planned)

    def set_hours_utilized(self, hours):
        if round(hours, 1) != self.hours_utilized:
            self.hours_utilized = round(hours, 1) #TODO: czy to jest dobrze liczone ?
            self.signal_hours_utilized_changed.emit(self.hours_utilized)

    def set_released(self, state):
        self.released = int(state)
        self._update_status()
        self.signal_released_changed.emit(self.released)

    # TODO: to jest wykorzystywane teraz do wczytania stack tasków
    def load_task_names_by_responsible_login(self, login):
        self.ids.clear()
        self.names.clear()

        cmd = '''SELECT 
                    id, 
                    name,
                    true_delivery
                FROM 
                    tasks 
                WHERE 
                    responsible_login = ?'''
        param = [login, ]
        try:
            self.db.execute(cmd, param)
            for element in self.db.fetchall():
                self.ids.append(element[0])
                self.names.append(element[1])
                if element[2] is None:
                    self.statuses.append(0)
                else:
                    self.statuses.append(1)
        except Exception as err:
           self.logger.exception(f'FAILED loading task names by responsible login.\nError occured: ')

    def load_data_basic(self):
        if self.id in self.ram_tasks.keys():
            print('(TASK) loading from RAM')
            task_data = self.ram_tasks[self.id]
            self.name = task_data.name
            self.lock = task_data.lock
            self.locking = task_data.locking
            self.transaction_id = task_data.transaction_id
            self.discipline = task_data.discipline
            self.activity = task_data.activity
            self.creator = task_data.creator
            self.responsible = task_data.responsible
            self.planned_start = task_data.planned_start
            self.planned_end = task_data.planned_end
            self.planned_delivery = task_data.planned_delivery
            self.true_start = task_data.true_start
            self.true_end = task_data.true_end
            self.true_delivery = task_data.true_delivery
            self.hours_planned = task_data.hours_planned
            self.hours_utilized = task_data.hours_utilized
            self.deliverable = task_data.deliverable
            self.checker_required = task_data.checker_required
            self.released = task_data.released
            self.editable = task_data.editable
            self.checks = task_data.checks
            self.reworks = task_data.reworks
            self._update_status()
        else:
            task_data = Task(db=self.db)
            if self.id is not None and self.loaded_basic is False:
                print("(TASK) loading data - basics [id={}]".format(self.id))
                cmd = ('''SELECT
                            name,
                            lock,
                            locking_login,
                            transaction_id,
                            discipline_id,
                            activity_id,
                            creator_login,
                            responsible_login,
                            planned_start,
                            planned_end,
                            planned_delivery,
                            true_start,
                            true_end,
                            true_delivery,
                            hours_planned,
                            hours_utilized,
                            deliverable,
                            checking_required,
                            released,
                            editable
                        FROM
                            tasks
                        WHERE
                            id = ?''')

                param = [self.id, ]
                try:
                    self.db.execute(cmd, param)
                except Exception:
                    self.logger.exception(f'FAILED loading basic data: ')
                else:

                    try:
                        sql_res = self.db.fetchone()
                    except Exception:
                        self.logger.exception(f'FAILED fetching data')
                    else:
                        self.name = sql_res[0]
                        self.lock = sql_res[1]
                        self.locking.login = sql_res[2]
                        self.transaction_id = sql_res[3]
                        self.discipline.id = sql_res[4]
                        self.discipline.load_data()
                        self.activity.id = sql_res[5]
                        self.activity.load_data()
                        self.creator.login = sql_res[6]
                        self.creator.load_data()
                        self.responsible.login = sql_res[7]
                        self.responsible.load_data()
                        self.planned_start = self.convert_to_date_class(sql_res[8])
                        self.planned_end = self.convert_to_date_class(sql_res[9])
                        self.planned_delivery = self.convert_to_date_class(sql_res[10])
                        self.true_start = self.convert_to_date_class(sql_res[11])
                        self.true_end = self.convert_to_date_class(sql_res[12])
                        self.true_delivery = self.convert_to_date_class(sql_res[13])
                        self.hours_planned = sql_res[14]
                        self.hours_utilized = sql_res[15]
                        self.deliverable = sql_res[16]
                        self.checker_required = int(sql_res[17])
                        self.released = sql_res[18]
                        self.editable = sql_res[19]

                        self.loaded_basic = True
                        self._load_check()
                        self._load_reworks()
                        self._update_status()

                        if self.released == 1:
                            task_data.name = self.name
                            task_data.lock = self.lock
                            task_data.locking = self.locking
                            task_data.transaction_id = self.transaction_id
                            task_data.discipline = self.discipline
                            task_data.activity = self.activity
                            task_data.creator = self.creator
                            task_data.responsible = self.responsible
                            task_data.planned_start = self.planned_start
                            task_data.planned_end = self.planned_end
                            task_data.planned_delivery = self.planned_delivery
                            task_data.true_start = self.true_start
                            task_data.true_end = self.true_end
                            task_data.true_delivery = self.true_delivery
                            task_data.hours_planned = self.hours_planned
                            task_data.hours_utilized = self.hours_utilized
                            task_data.deliverable = self.deliverable
                            task_data.checker_required = self.checker_required
                            task_data.released = self.released
                            task_data.checks = self.checks
                            task_data.reworks = self.reworks
                            task_data.editable = self.editable
                            task_data.status = self.status

                            self.ram_tasks[self.id] = task_data

    def load_task_comments(self):
        self._load_comments()

    def has_childs(self):
        cmd = '''SELECT
                       id
                   FROM
                       checks
                   WHERE
                       task_id=?'''
        param = [self.id]
        try:
            self.db.execute(cmd, param)
        except Exception:
            self.logger.exception(f'FAILED loading checks for task:')
        else:
            try:
                sql_res = self.db.fetchall()
            except Exception:
                self.logger.exception(f'FAILED to fetch checks for task:')
            else:
                if len(sql_res) == 0:
                    return False
                else:
                    return True

    def _load_check(self):
        self.checks=[]
        cmd = '''SELECT
                    id
                FROM
                    checks
                WHERE
                    task_id=?'''
        param = [self.id]
        try:
            self.db.execute(cmd, param)
        except Exception:
            self.logger.exception(f'FAILED loading checks for task:')
        else:
            try:
                sql_res = self.db.fetchall()
            except Exception:
                self.logger.exception(f'FAILED to fetch checks for task:')
            else:
                for element in sql_res:
                    check = Check(db=self.db)
                    check.set_id(element[0])
                    check.load_data()
                    self.checks.append(check)
                    self.check = check

    def _load_reworks(self):
        self.reworks = []
        cmd = '''SELECT
                       id
                   FROM
                       reworks
                   WHERE
                       task_id=?'''
        param = [self.id]
        try:
            self.db.execute(cmd, param)
        except Exception:
            self.logger.exception(f'FAILED loading checks for task: ')
        else:
            try:
                sql_res = self.db.fetchall()
            except Exception:
                self.logger.exception(f'FAILED fetching')
            else:
                for element in sql_res:
                    rework = Rework(db=self.db)
                    rework.set_id(element[0])
                    rework.load_data()
                    self.reworks.append(rework)
                    self.rework = rework

    def _load_comments(self):
        self.comments = []
        cmd = '''SELECT
                    id
                FROM
                    comments
                WHERE
                    task_id=?'''
        param = [self.id]
        try:
            self.db.execute(cmd, param)
        except Exception:
            self.logger.exception(f'FAILED loading comments for task.\nError occured: ')
        else:
            sql_res = self.db.fetchall()
            for element in sql_res:
                comment = Comment(db=self.db)
                comment.set_id(element[0])
                self.comments.append(comment)

    def _load_transaction_info(self):
        if self.transaction.id != self.transaction_id:
            self.transaction.set_id(self.transaction_id)
            self.transaction.load_data_basic()
        else:
            if self.transaction.loaded_basic is False:
                self.transaction.load_data_basic()

    def reassign_to_user(self, user):
        cmd = '''UPDATE
                    tasks
                SET
                    responsible_login=?
                WHERE
                    id= ?'''
        param = [user.login, self.id]
        try:
            self.db.execute(cmd, param)
            self.db.commit()
        except Exception:
            self.logger.exception(f'REASSIGMENT FAILED! Exception occured\n')
            self.signal_message(f'Task reassigmend FAILED! Please try again')
            self.signal_updated.emit(False)
        else:
            self.signal_message.emit("Task sucessfully reassigned")
            self.signal_updated.emit(True)

    def update_data(self):
        if self.ready_to_save():
            cmd = '''UPDATE
                        tasks
                    SET
                        name = ?,
                        transaction_id = ?,
                        discipline_id = ?,
                        activity_id = ?,
                        creator_login = ?,
                        responsible_login = ?,
                        planned_start = ?,
                        planned_end = ?,
                        planned_delivery = ?,
                        true_start = ?,
                        true_end = ?,
                        true_delivery = ?,
                        hours_planned = ?,
                        hours_utilized = ?,
                        deliverable = ?,
                        checking_required = ?,
                        released = ?
                    WHERE
                        id = ?
                    '''

            param = [self.name,
                     self.transaction_id,
                     self.discipline.id,
                     self.activity.id,
                     self.creator.login,
                     self.responsible.login,
                     self.convert_to_string_class(self.planned_start),
                     self.convert_to_string_class(self.planned_end),
                     self.convert_to_string_class(self.planned_delivery),
                     self.convert_to_string_class(self.true_start),
                     self.convert_to_string_class(self.true_end),
                     self.convert_to_string_class(self.true_delivery),
                     self.hours_planned,
                     self.hours_utilized,
                     self.deliverable,
                     self.checker_required,
                     self.released,
                     self.id]
            try:
                self.db.execute(cmd, param)
                self.db.commit()
            except Exception :
                self.logger.exception(f'UPDATE FAILED!')
                self.signal_message(f'UPDATE FAILED! Check connection and try again')
                self.signal_updated.emit(False)
            else:
                self.signal_message.emit("Task updated successfully")
                self.signal_updated.emit(True)

    def register_data(self):
        creator_login = os.getlogin().upper()
        if self.ready_to_save():
            cmd = "INSERT INTO tasks VALUES(NULL, ?, 0, NULL, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)"
            param = [self.name,
                     self.transaction_id,
                     self.discipline.id,
                     self.activity.id,
                     creator_login,
                     self.responsible.login,
                     self.planned_start,
                     self.planned_end,
                     self.planned_delivery,
                     self.true_start,
                     self.true_end,
                     self.true_delivery,
                     self.hours_planned,
                     self.hours_utilized,
                     self.deliverable,
                     self.checker_required,
                     self.released,
                     self.editable]
            try:
                self.db.execute(cmd, param)
                self.db.commit()
            except Exception:
                self.logger.exception(f'SAVE FAILED!')
                self.signal_message.emit(f'SAVE FAILED! Please try again')
                self.signal_created.emit(False)
            else:
                self.signal_message.emit("Task saved successfully")
                self.signal_created.emit(True)

    def delete_data(self):
        if self.id is not None:
            cmd = f'DELETE FROM tasks WHERE id=?'
            param = [self.id, ]
            try:
                self.db.execute(cmd, param)
                self.db.commit()
            except Exception:
                self.signal_deleted.emit(False)
                self.logger.exception(f'Error while trying to execute delete command on database')
            else:
                self.signal_deleted.emit(True)

    def ready_to_save(self):

        if self.transaction_id is None:
            self.signal_message.emit("Transaction can't be None")
            return False
        else:
            self._load_transaction_info()

            if (self.transaction.planned_start - self.planned_start).days > 0:
                self.signal_message.emit(f"Task Planned Start ({self.planned_start}) can't be before Transaction Planned Start({self.transaction.planned_start})")
                return False

            unresonable_error_causing = self.transaction.deliverable

            if self.transaction.deliverable == 1:
                if (self.transaction.planned_delivery - self.planned_end).days < 0:
                    self.signal_message.emit(f"Task Planned End ({self.planned_end}) can't be after Transaction Planned Delivery ({self.transaction.planned_delivery})")
                    return False
            else:
                if self.deliverable == 1:
                    if (self.transaction.planned_end - self.planned_delivery).days < 0:
                        self.signal_message.emit(f"Task Planned Delivery ({self.planned_delivery}) can't be after Transaction Planned End({self.transaction.planned_end})")
                        return False
                else:
                    if (self.transaction.planned_end - self.planned_end).days < 0:
                        self.signal_message.emit(f"Task Planned End ({self.planned_end}) can't be after Transaction Planned End ({self.transaction.planned_end})")
                        return False

        if self.discipline.id is None:
            self.signal_message.emit("Discipline can't be None")
            return False

        if self.activity.id is None:
            self.signal_message.emit("Activity Type can't be None")
            return False

        if self.name is None or self.name is "":
            self.signal_message.emit("Name can't be empty")
            return False

        if self.responsible.login is None:
            self.signal_message.emit("Responsible person can't be None")

            return False

        if self.planned_start is None:
            self.signal_message.emit("Planned Start date can't be None")
            return False

        if self.planned_end is None:
            self.signal_message.emit("Planned End date can't be None")
            return False

        if self.planned_start is not None and self.planned_end is not None:
            if (self.planned_end - self.planned_start).days < 0:
                self.signal_message.emit("Planned End can't be before Planned Start")
                return False

        if self.deliverable is None:
            self.signal_message.emit(f'Select YES/NO for deliverable option')
            return False

        if self.deliverable == 1:
            if self.planned_delivery is None:
                self.signal_message.emit("If task is Deliverable then Planned Delivery date can't be None")
                return False

            if self.planned_end is not None and self.planned_delivery is not None:
                if (self.planned_delivery - self.planned_end).days < 0:
                    self.signal_message.emit("Planned Delivery can't be before Planned End")
                    return False

        if self.checker_required is None:
            self.signal_message.emit("Select YES/NO option for checking requirement ")
            return False

        if self.lock == 1 and self.locking.login != os.getlogin().upper():
            self.signal_message.emit(f"Current task is locked for editing by {self.locking.full_name}")
            return False

        self.signal_message.emit("Task ready to be saved")
        return True

#TODO: To be developed later
    def ready_to_save2(self):
        rts_signal = {"ready": True}

        if self.transaction_id is None:
            rts_signal["ready"] = False
            rts_signal["transaction"] = False
        else:
            rts_signal["transaction"] = True

        if self.discipline.id is None:
            rts_signal["ready"] = False
            rts_signal["discipline"] = False
        else:
            rts_signal["discipline"] = True

        if self.activity.id is None:
            rts_signal["ready"] = False
            rts_signal["activity"] = False
        else:
            rts_signal["activity"] = True

        if self.name is None or self.name is "":
            rts_signal["ready"] = False
            rts_signal["name"] = False
        else:
            rts_signal["name"] = True

        if self.responsible.login is None:
            rts_signal["ready"] = False
            rts_signal["responsible"] = False
        else:
            rts_signal["responsible"] = True

        if self.planned_start is None:
            rts_signal["ready"] = False
            rts_signal["planned_start"] = False
        else:
            rts_signal["planned_start"] = True

        if self.checker_required is None:
            rts_signal["ready"] = False
            rts_signal["checking_required"] = False
        else:
            rts_signal["checking_required"] = True

        return rts_signal


class Check(DateManager, QObject):
    signal_created = pyqtSignal(bool)
    signal_updated = pyqtSignal(bool)
    signal_id_changed = pyqtSignal(int)
    signal_status_changed = pyqtSignal(int)
    signal_planned_start_changed = pyqtSignal(object)
    signal_planned_end_changed = pyqtSignal(object)
    signal_true_start_changed = pyqtSignal(object)
    signal_true_end_changed = pyqtSignal(object)
    signal_hours_planned_changed = pyqtSignal(float)
    signal_hours_utilized_changed = pyqtSignal(float)
    signal_availability_changed = pyqtSignal(float)
    signal_rework_changed = pyqtSignal(int)
    signal_released_changed = pyqtSignal(int)
    signal_message = pyqtSignal(str)

    def __init__(self, db=None):
        QObject.__init__(self)
        DateManager.__init__(self)
        self.logger = logging.getLogger(self.__class__.__name__)
        if db:
            self.db = db
        else:
            self.logger.warning(f' used NEW HANDLER for connection')
            self.db = ConnectionDB()

        # SQL TABLE
        self.id = None
        self.task_id = None
        self.creator = Employee(db=self.db)
        self.responsible = Employee(db=self.db)
        self.planned_start = None
        self.planned_end = None
        self.true_start = None
        self.true_end = None
        self.hours_planned = 0
        self.hours_utilized = 0
        self.availability = 0
        self.rework = 0
        self.released = 0

        # LISTS
        self.ids = []

        # OTHERS
        self.status = 0
        self.loaded = False

    def _update_status(self):

        if self.released == 1:
            if self.rework == 0:
                self._set_status(status.code["RELEASED"])
                return
            if self.rework == 1:
                self._set_status(status.code["RETURNED"])
                return

        if self.true_end is not None:
            self._set_status(status.code["FINISHED"])
            return

        if self.true_start is not None:
            if self.planned_end < date.today():
                self._set_status(status.code["LATE"])
                return
            else:
                self._set_status(status.code["ONGOING"])
                return

        if self.planned_start is not None and self.planned_end is not None:
            self._set_status(status.code["SCHEDULED"])
            return

        if self.responsible.login is not None:
            if self.responsible.login[:11] != "STACK_CHECK":
                self._set_status(status.code["ASSIGNED"])
                return
            else:
                self._set_status(status.code["READY TO POP"])
                return

        self._set_status(status.code["ERROR"])

    def _set_status(self, code):
        self.status = code
        self.signal_status_changed.emit(self.status)

    def reset(self):
        self.id = None
        self.task_id = None
        self.creator.reset()
        self.responsible.reset()
        self.planned_start = None
        self.planned_end = None
        self.true_start = None
        self.true_end = None
        self.hours_planned = 0
        self.hours_utilized = 0
        self.availability = 0
        self.rework = 0
        self.released = 0
        self.status = 0
        self.loaded = False

    def set_id(self, id):
        self.reset()
        self.id = id
        self.load_data()
        self.signal_id_changed.emit(self.id)

    def set_id_by_index(self, index):
        self.set_id(self.ids[index])

    def set_task_id(self, task_id):
        self.task_id = task_id

    def set_responsible_login(self, login):
        self.responsible.set_login(login)
        self._update_status()

    def set_planned_start(self, date_to_set):
        if date_to_set is None:
            self.planned_start = None
            self._update_status()
            self.signal_planned_start_changed.emit(self.planned_start)
        else:
            if isinstance(date_to_set, date):
                if date_to_set != self.planned_start:
                    self.planned_start = date_to_set
                    self._update_status()
                    self.signal_planned_start_changed.emit(self.planned_start)
            else:
                self.logger.critical(f'setting planned_start caused ERROR.\ngiven date format is not date instance {date_to_set}')
                sys.exit(3)

    def set_planned_end(self, date_to_set):
        if date_to_set is None:
            self.planned_end = None
            self._update_status()
            self.signal_planned_end_changed.emit(self.planned_end)
        else:
            if isinstance(date_to_set, date):
                if date_to_set != self.planned_end:
                    self.planned_end = date_to_set
                    self._update_status()
                    self.signal_planned_end_changed.emit(self.planned_end)
            else:
                self.logger.critical(f'setting planned_end caused ERROR.\ngiven date format is not date instance {date_to_set}')
                sys.exit(3)

    def set_true_start(self, date_to_set):
        if date_to_set is None:
            if self.true_start is not None:
                self.true_start = None
                self._update_status()
                self.signal_true_start_changed.emit(self.true_start)
        else:
            if isinstance(date_to_set, date):
                if date_to_set != self.true_start:
                    self.true_start = date_to_set
                    self._update_status()
                    self.signal_true_start_changed.emit(self.true_start)
            else:
                self.logger.critical(f'setting true_start caused ERROR.\ngiven date format is not date instance {date_to_set}')
                sys.exit(3)

    def set_true_end(self, date_to_set):
        if date_to_set is None:
            if self.true_end is not None:
                self.true_end = None
                self._update_status()
                self.signal_true_end_changed.emit(self.true_end)
        else:
            if isinstance(date_to_set, date):
                if date_to_set != self.true_end:
                    self.true_end = date_to_set
                    self._update_status()
                    self.signal_true_end_changed.emit(self.true_end)
            else:
                self.logger.critical(
                    f'setting true_end caused ERROR.\ngiven date format is not date instance {date_to_set}')
                sys.exit(3)

    def set_avability(self, avability):
        if avability != self.avability:
            self.logger.debug(f'setting avability. was {self.avability} is {int(avability)}')
            self.avability = round(avability, 2)
            self.signal_availability_changed.emit(self.avability)

    def set_hours_planned(self, hours):
        if round(hours, 1) != self.hours_planned:
            self.logger.debug(f'setting hours_planned. was {self.hours_planned} is {(hours)}')
            self.hours_planned = round(hours, 1)
            self.signal_hours_planned_changed.emit(self.hours_planned)

    def set_hours_utilized(self, hours):
        if round(hours, 1) != self.hours_utilized:
            self.logger.debug(f'setting hours_utilized. was {self.hours_planned} is {hours}')
            self.hours_utilized = round(hours, 1)
            self.signal_hours_utilized_changed.emit(self.hours_utilized)

    def set_rework(self, state):
        self.rework = int(state)
        self._update_status()
        self.signal_rework_changed.emit(self.rework)

    def set_released(self, state):
        self.released = int(state)
        self._update_status()
        self.signal_released_changed.emit(self.released)

    def load_data(self):
        if self.id is not None and self.loaded is False:
            print('(CHECK) loading data [id={}]'.format(self.id))
            cmd = ('''SELECT
                        task_id,
                        creator_login,
                        responsible_login,
                        planned_start,
                        planned_end,
                        true_start,
                        true_end,
                        hours_planned,
                        hours_utilized,
                        rework,
                        released
                    FROM
                        checks
                    WHERE
                        id = ?''')
            param = [self.id, ]
            try:
                self.db.execute(cmd, param)
            except Exception:
                self.logger.exception(f'FAILED loading data.\nError occured: ')
            else:
                sql_res = self.db.fetchone()
                self.set_task_id(sql_res[0])
                self.creator.set_login(sql_res[1])
                self.responsible.set_login(sql_res[2])
                self.set_planned_start(self.convert_to_date_class(sql_res[3]))
                self.set_planned_end(self.convert_to_date_class(sql_res[4]))
                self.set_true_start(self.convert_to_date_class(sql_res[5]))
                self.set_true_end(self.convert_to_date_class(sql_res[6]))
                self.set_hours_planned(sql_res[7])
                self.set_hours_utilized(sql_res[8])
                self.set_rework(sql_res[9])
                self.set_released(sql_res[10])

                self.loaded = True
                self._update_status()

    def register_data(self):
        self.logger.info(f'trying to save check')
        self.creator.set_login(os.getlogin())

        cmd = "INSERT INTO checks VALUES(NULL,?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)"
        param = [self.task_id,
                 self.creator.login,
                 self.responsible.login,
                 self.planned_start,
                 self.planned_end,
                 self.true_start,
                 self.true_end,
                 self.hours_planned,
                 self.hours_utilized,
                 self.rework,
                 self.released]
        try:
            self.db.execute(cmd, param)
            self.db.commit()
        except Exception:
            self.logger.exception(f'CREATION FAILED! Exception occured\n')
            self.signal_created.emit(False)
        else:
            self.logger.info(f'check created successfully with id#{self.db.cursor.lastrowid}')
            self.id = self.db.cursor.lastrowid
            self.signal_created.emit(True)

    def update_data(self):
        cmd = '''UPDATE
                    checks
                SET
                    task_id = ?,
                    creator_login = ?,
                    responsible_login = ?,
                    planned_start = ?,
                    planned_end = ?,
                    true_start = ?,
                    true_end = ?,
                    hours_planned = ?,
                    hours_utilized = ?,
                    rework = ?,
                    released = ?
                WHERE
                    id = ?
                '''
        param = [self.task_id,
                 self.creator.login,
                 self.responsible.login,
                 self.convert_to_string_class(self.planned_start),
                 self.convert_to_string_class(self.planned_end),
                 self.convert_to_string_class(self.true_start),
                 self.convert_to_string_class(self.true_end),
                 self.hours_planned,
                 self.hours_utilized,
                 self.rework,
                 self.released,
                 self.id]
        try:
            self.db.execute(cmd, param)
            self.db.commit()
        except Exception:
            self.logger.exception(f'UPDATE FAILED! Exception occured\n')
            self.signal_message(f'UPDATE FAILED! Please try again')
            self.signal_updated.emit(False)
        else:
            self.logger.info(f'task saved successfully. id {self.id}')
            self.signal_message.emit("Task saved successfully")
            self.signal_updated.emit(True)


class Rework(QObject, DateManager):
    signal_created = pyqtSignal(bool)
    signal_updated = pyqtSignal(bool)
    signal_id_changed = pyqtSignal(int)
    signal_status_changed = pyqtSignal(int)
    signal_planned_start_changed = pyqtSignal(object)
    signal_planned_end_changed = pyqtSignal(object)
    signal_true_start_changed = pyqtSignal(object)
    signal_true_end_changed = pyqtSignal(object)
    signal_hours_planned_changed = pyqtSignal(float)
    signal_hours_utilized_changed = pyqtSignal(float)
    signal_availability_changed = pyqtSignal(float)
    signal_released_changed = pyqtSignal(int)
    signal_message = pyqtSignal(str)

    def __init__(self, db=None):
        self.logger = logging.getLogger(self.__class__.__name__)
        self.logger.debug(f'creating  object')
        QObject.__init__(self)
        DateManager.__init__(self)

        if db:
            self.db = db
        else:
            self.logger.warning(f' used NEW HANDLER for connection')
            self.db = ConnectionDB()

        # SQL TABLE
        self.id = None
        self.task_id = None
        self.creator = Employee(db=self.db)
        self.responsible = Employee(db=self.db)
        self.planned_start = None
        self.planned_end = None
        self.true_start = None
        self.true_end = None
        self.hours_planned = 0
        self.hours_utilized = 0
        self.availability = 0
        self.released = 0

        # LISTS
        self.ids = []

        # OTHERS
        self.status = 0
        self.loaded = False

        self._connect_signals()

    def _connect_signals(self):
        self.signal_id_changed.connect(self._update_status)
        self.responsible.signal_login_changed.connect(self._update_status)
        self.signal_planned_start_changed.connect(self._update_status)
        self.signal_planned_end_changed.connect(self._update_status)
        self.signal_true_start_changed.connect(self._update_status)
        self.signal_true_end_changed.connect(self._update_status)
        self.signal_released_changed.connect(self._update_status)

    def _set_status(self, code):
        self.status = code
        self.signal_status_changed.emit(self.status)

    def _update_status(self):

        if self.released == 1:
            self._set_status(status.code["RELEASED"])
            return

        if self.true_end is not None:
            self._set_status(status.code["FINISHED"])
            return

        if self.true_start is not None:
            if self.planned_end < date.today():
                self._set_status(status.code["LATE"])
                return
            else:
                self._set_status(status.code["ONGOING"])
                return

        if self.planned_start is not None and self.planned_end is not None:
            self._set_status(status.code["SCHEDULED"])
            return

        if self.responsible.login is not None:
            self._set_status(status.code["ASSIGNED"])
            return

        self._set_status(status.code["ERROR"])

    def reset(self):
        self.id = None
        self.task_id = None
        self.creator.reset()
        self.responsible.reset()
        self.planned_start = None
        self.planned_end = None
        self.true_start = None
        self.true_end = None
        self.hours_planned = 0
        self.hours_utilized = 0
        self.availability = 0
        self.released = 0
        self.status = 0
        self.loaded = False

    def set_id(self, id):
        self.reset()
        self.id = id
        self.load_data()
        self.signal_id_changed.emit(self.id)

    def set_id_by_index(self, index):
        self.set_id(self.ids[index])

    def set_task_id(self, task_id):
        self.task_id = task_id

    def set_planned_start(self, date_to_set):
        if date_to_set is None:
            self.logger.debug(f'setting planned_start. was {self.planned_start} is {date_to_set}')
            self.planned_start = None
            self.signal_planned_start_changed.emit(None)
        else:
            if isinstance(date_to_set, date):
                if date_to_set != self.planned_start:
                    self.logger.debug(f'setting planned_start. was {self.planned_start} is {date_to_set}')
                    self.planned_start = date_to_set
                    self.signal_planned_start_changed.emit(self.planned_start)
            else:
                self.logger.critical(f'setting planned_start caused ERROR.\ngiven date format is not date instance {date_to_set}')
                sys.exit(3)

    def set_planned_end(self, date_to_set):
        if date_to_set is None:
            self.logger.debug(f'setting planned_end. was {self.planned_end} is {date_to_set}')
            self.planned_end = None
            self.signal_planned_end_changed.emit(None)
        else:
            if isinstance(date_to_set, date):
                if date_to_set != self.planned_end:
                    self.logger.debug(f'setting planned_end. was {self.planned_end} is {date_to_set}')
                    self.planned_end = date_to_set
                    self.signal_planned_end_changed.emit(self.planned_end)
            else:
                self.logger.critical(f'setting planned_end caused ERROR.\ngiven date format is not date instance {date_to_set}')
                sys.exit(3)

    def set_true_start(self, date_to_set):
        if date_to_set is None:
            if self.true_start is not None:
                self.logger.debug(f'setting true_start. was {self.true_start} is {date_to_set}')
                self.true_start = None
                self.signal_true_start_changed.emit(self.true_start)
        else:
            if isinstance(date_to_set, date):
                if date_to_set != self.true_start:
                    self.logger.debug(f'setting true_start. was {self.true_start} is {date_to_set}')
                    self.true_start = date_to_set
                    self.signal_true_start_changed.emit(self.true_start)
            else:
                self.logger.critical(f'setting true_start caused ERROR.\ngiven date format is not date instance {date_to_set}')
                sys.exit(3)

    def set_true_end(self, date_to_set):
        if date_to_set is None:
            if self.true_end is not None:
                self.logger.debug(f'setting true_END. was {self.true_end} is {date_to_set}')
                self.true_end = None
                self.signal_true_end_changed.emit(self.true_end)
        else:
            if isinstance(date_to_set, date):
                if date_to_set != self.true_end:
                    self.logger.debug(f'setting true_end. was {self.true_end} is {date_to_set}')
                    self.true_end = date_to_set
                    self.signal_true_end_changed.emit(self.true_end)
            else:
                self.logger.critical(
                    f'setting true_end caused ERROR.\ngiven date format is not date instance {date_to_set}')
                sys.exit(3)

    def set_hours_planned(self, hours):
        if round(hours, 1) != self.hours_planned:
            self.logger.debug(f'setting hours_planned. was {self.hours_planned} is {(hours)}')
            self.hours_planned = round(hours, 1)
            self.signal_hours_planned_changed.emit(self.hours_planned)

    def set_hours_utilized(self, hours):
        if round(hours, 1) != self.hours_utilized:
            self.logger.debug(f'setting hours_utilized. was {self.hours_planned} is {hours}')
            self.hours_utilized = round(hours, 1)
            self.signal_hours_utilized_changed.emit(self.hours_utilized)

    def set_rework(self, state):
        self.rework = int(state)
        self.signal_rework_changed.emit(self.rework)

    def set_released(self, state):
        self.released = int(state)
        self.signal_released_changed.emit(self.released)

    def load_data(self):
        self.logger.debug(f'requested load_data')
        if self.id is not None and self.loaded is False:
            print('(REWORK) loading data')
            cmd = ('''SELECT
                        task_id,
                        creator_login,
                        responsible_login,
                        planned_start,
                        planned_end,
                        true_start,
                        true_end,
                        hours_planned,
                        hours_utilized,
                        released
                    FROM
                        reworks
                    WHERE
                        id = ?''')
            param = [self.id, ]
            try:
                self.db.execute(cmd, param)
            except Exception:
                self.logger.exception(f'FAILED loading data.\nError occured: ')
            else:
                sql_res = self.db.fetchone()
                self.set_task_id(sql_res[0])
                self.creator.set_login(sql_res[1])
                self.creator.load_data()
                self.responsible.set_login(sql_res[2])
                self.responsible.load_data()
                self.set_planned_start(self.convert_to_date_class(sql_res[3]))
                self.set_planned_end(self.convert_to_date_class(sql_res[4]))
                self.set_true_start(self.convert_to_date_class(sql_res[5]))
                self.set_true_end(self.convert_to_date_class(sql_res[6]))
                self.set_hours_planned(sql_res[7])
                self.set_hours_utilized(sql_res[8])
                self.set_released(sql_res[9])

                self.loaded = True
                self._update_status()

    def register_data(self):
        self.logger.info(f'trying to save check')
        self.creator.set_login(os.getlogin())

        cmd = "INSERT INTO reworks VALUES(NULL,?, ?, ?, ?, ?, ?, ?, ?, ?, ?)"
        param = [self.task_id,
                 self.creator.login,
                 self.responsible.login,
                 self.planned_start,
                 self.planned_end,
                 self.true_start,
                 self.true_end,
                 self.hours_planned,
                 self.hours_utilized,
                 self.released]
        try:
            self.db.execute(cmd, param)
            self.db.commit()
        except Exception:
            self.logger.exception(f'CREATION FAILED! Exception occured\n')
            self.signal_created.emit(False)
        else:
            self.logger.info(f'check created successfully with id#{self.db.cursor.lastrowid}')
            self.id = self.db.cursor.lastrowid
            self.signal_created.emit(True)

    def update_data(self):
        cmd = '''UPDATE
                    reworks
                SET
                    task_id = ?,
                    creator_login = ?,
                    responsible_login = ?,
                    planned_start = ?,
                    planned_end = ?,
                    true_start = ?,
                    true_end = ?,
                    hours_planned = ?,
                    hours_utilized = ?,
                    released = ?
                WHERE
                    id = ?
                '''
        param = [self.task_id,
                 self.creator.login,
                 self.responsible.login,
                 self.convert_to_string_class(self.planned_start),
                 self.convert_to_string_class(self.planned_end),
                 self.convert_to_string_class(self.true_start),
                 self.convert_to_string_class(self.true_end),
                 self.hours_planned,
                 self.hours_utilized,
                 self.released,
                 self.id]
        try:
            self.db.execute(cmd, param)
            self.db.commit()
        except Exception:
            self.logger.exception(f'UPDATE FAILED! Exception occured\n')
            self.signal_message(f'UPDATE FAILED! Please try again')
            self.signal_updated.emit(False)
        else:
            self.logger.info(f'task saved successfully. id {self.id}')
            self.signal_message.emit("Task saved successfully")
            self.signal_updated.emit(True)


class Comment(QObject, DateManager):
    signal_message = pyqtSignal(str)
    signal_id_changed = pyqtSignal(int)
    ram_comments = dict()

    def __init__(self, db=None):
        QObject.__init__(self)
        DateManager.__init__(self)
        if db:
            self.db = db
        else:
            self.db = ConnectionDB()

        # SQL TABLE
        self.id = None
        self.source = None
        self.transaction_id = None
        self.task_id = None
        self.description = None
        self.publishing_date = datetime.now().strftime("%Y-%m-%d %H:%M")
        self.publisher = Employee(db=self.db)

    def _reset(self):
        self.id = None
        self.source = None
        self.transaction_id = None
        self.task_id = None
        self.description = None
        self.publishing_date = None
        self.publisher.set_login(None)

    def _load_data(self):
        if self.id in self.ram_comments.keys():
            print('(COMMENT) loading data from RAM')
            self.source = self.ram_comments[self.id].source
            self.description = self.ram_comments[self.id].description
            self.publishing_date = self.ram_comments[self.id].publishing_date
            self.publisher = self.ram_comments[self.id].publisher
        else:
            ram_comment = Comment(db=self.db)
            if self.id is not None:
                print('(COMMENT) loading data from DATABASE')
                cmd = '''SELECT
                            source,
                            transaction_id,
                            task_id,
                            description,
                            publishing_date,
                            publisher_login,
                            include
                        FROM
                            comments
                        WHERE
                            id=?'''
                param = [self.id, ]
                try:
                    self.db.execute(cmd, param)
                except Exception:
                    self.logger.exception(f'Error while trying to execute command on database')
                else:
                    try:
                        sql_res = self.db.fetchone()
                    except Exception:
                        self.logger.exception(f'Error while trying to fetching data from database')
                    else:
                        self.source = sql_res[0]
                        self.transaction_id = sql_res[1]
                        self.task_id = sql_res[2]
                        self.description = sql_res[3]
                        self.publishing_date = sql_res[4]
                        self.publisher.set_login(sql_res[5])
                        self.publisher.load_data()

                        ram_comment.source = self.source
                        ram_comment.description = self.description
                        ram_comment.publishing_date = self.publishing_date
                        ram_comment.publisher = self.publisher

                        self.ram_comments[self.id] = ram_comment

    def set_id(self, id):
        self._reset()
        self.id = id
        self._load_data()
        # self.signal_id_changed.emit(self.id)

    def set_description(self, description):
        if description != self.description:
            self.description = str(description)

    def set_source(self, source):
        if source != self.source:
            self.source = str(source)

    def set_task_id(self, task_id):
        if task_id != self.task_id:
            self.task_id = task_id

    def set_transaction_id(self, transaction_id):
        if transaction_id != self.transaction_id:
            self.transaction_id = transaction_id

    def set_publisher_login(self, login):
        if login != self.publisher.login:
            self.publisher.set_login(login)

    def register_data(self):
        if (self.description and self.task_id) or (self.description and self.transaction_id):
            cmd = '''INSERT INTO comments VALUES (NULL, ?, ?, ?, ?, ?, ?, NULL)'''
            param = [self.source,
                     self.transaction_id,
                     self.task_id,
                     self.description,
                     self.publishing_date,
                     self.publisher.login]
            try:
                self.db.execute(cmd, param)
                self.db.commit()
            except Exception:
                self.logger.exception(f'Error while trying to save new comment')


class Loader(QObject, DateManager):

    def __init__(self, user):
        DateManager.__init__(self)
        QObject.__init__(self)
        self.logger = logging.getLogger(self.__class__.__name__)
        self.me = user

        self.transactions = []

        # FILTERS
        self.responsible_only = False
        self.customer_id = None
        self.filter_show_released_task = False
        self.filter_show_released_transactions = False
        self.filter_show_released_checks = False

    def transactions_where_user_is_responsible(self):
        self.transactions = []
        cmd = '''SELECT
                       id
                   FROM
                       transactions
                   WHERE
                       responsible_login=?'''
        param = [self.me.login, ]
        if self.filter_show_released_transactions is True:
            cmd += 'AND (released = 1 OR released = 0)'
        else:
            cmd += 'AND released = 0'

        try:
            self.me.db.execute(cmd, param)
        except Exception:
            self.logger.exception(f'Error while sending cmd and param database')
        else:
            try:
                transaction_ids = self.me.db.fetchall()
            except Exception:
                self.logger.exception(f'Error while receiving data from database')
            else:
                for id in transaction_ids:
                    transaction = Transaction(db=self.me.db)
                    transaction.filter_show_released_task = self.filter_show_released_task
                    transaction.filter_show_released_checks = self.filter_show_released_checks
                    transaction.set_id(id[0])
                    transaction.load_data_basic()
                    transaction.load_data_childs()
                    self.transactions.append(transaction)

    def transactions_where_user_has_actions(self):
        self.transactions = []
        transaction_ids = []

        if self.filter_show_released_task is True:
            filter_task_cmd = '(released = 1 OR released = 0)'
        else:
            filter_task_cmd = 'released = 0'

        cmd = '''SELECT
                       transaction_id
                   FROM
                       tasks
                   WHERE
                       responsible_login=?
                   AND
                       {filter_task_cmd}'''.format(filter_task_cmd=filter_task_cmd)
        param = [self.me.login, ]

        try:
            self.me.db.execute(cmd, param)
        except Exception:
            self.logger.exception(f'Error while trying to send cmd to database')
        else:
            try:
                all_transactions = self.me.db.fetchall()
            except Exception:
                self.logger.exception(f'Error while trying to download data from database')
            else:
                for id in all_transactions:
                    if id[0] not in transaction_ids:
                        transaction_ids.append(id[0])

        if self.filter_show_released_checks is True:
            filter_check_cmd = '(released=1 OR released=0)'
        else:
            filter_check_cmd = 'released=0'

        cmd = '''SELECT
                       transaction_id
                   FROM
                       tasks
                   WHERE
                       id
                   IN 
                       (SELECT
                           task_id
                       FROM
                           checks
                       WHERE
                           responsible_login = ?
                       AND
                           {filter_check_cmd}
                       )'''.format(filter_check_cmd=filter_check_cmd)
        param = [self.me.login, ]

        try:
            self.me.db.execute(cmd, param)
        except Exception:
            self.logger.exception(f'Error while trying to execute command to database')
        else:
            try:
                all_transactions = self.me.db.fetchall()
            except Exception:
                self.logger.exception(f'Error while trying to fetching data from database')
            else:
                for id in all_transactions:
                    if id[0] not in transaction_ids:
                        transaction_ids.append(id[0])

        for id in transaction_ids:
            transaction = Transaction(db=self.me.db)
            transaction.filter_show_released_task = self.filter_show_released_task
            transaction.filter_show_released_checks = self.filter_show_released_checks
            transaction.set_id(id)
            transaction.load_data_basic()
            transaction.load_data_childs(employee_login=self.me.login)
            self.transactions.append(transaction)

    def transaction_for_customer(self):
        self.transactions = []
        cmd = '''SELECT 
                       id 
                   FROM 
                       transactions 
                   WHERE 
                       customer_id = ? '''
        param = [self.customer_id]

        if self.filter_show_released_transactions:
            pass
        else:
            cmd += 'AND released = 0'
        try:
            self.me.db.execute(cmd, param)
        except Exception:
            self.logger.exception(f'Error while sending cmd and param database')
        else:
            try:
                transaction_ids = self.me.db.fetchall()
            except Exception:
                self.logger.exception(f'Error while receiving data from database')
            else:
                for id in transaction_ids:
                    transaction = Transaction(db=self.me.db)
                    transaction.filter_show_released_checks = self.filter_show_released_checks
                    transaction.filter_show_released_task = self.filter_show_released_task
                    transaction.set_id(id[0])
                    transaction.load_data_basic()
                    transaction.load_data_childs()
                    self.transactions.append(transaction)
